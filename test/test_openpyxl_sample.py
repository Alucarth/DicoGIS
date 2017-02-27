from __future__ import print_function
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import Style, Font, Alignment

## styles
header = Style(font=Font(bold=True))
multilines = Style(alignment=Alignment(wrap_text=True))


wb = Workbook()
dest_filename = 'sample_workbook'
ws1 = wb.active
ws1.title = "range names"
for row in range(1, 40):
    ws1.append(range(600))

row1 = ws1.row_dimensions[1]
row1.style = header

wsprops = ws1.sheet_properties
wsprops.tabColor = "1072BA"
wsprops.filterMode = True

ws1.auto_filter.add_filter_column(1, {}, True)
ws1.auto_filter.add_filter_column(2, ["A1:A40"], True)


ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))
print(ws3['AA10'].value)

ws3.cell('A1').value = "Line 1\nLine 2\nLine 3"
ws3.cell('A1').style = multilines


ws4 = wb.create_sheet(title="Filtered data")

data = [
    ["Fruit", "Quantity"],
    ["Kiwi", 3],
    ["Grape", 15],
    ["Apple", 3],
    ["Peach", 3],
    ["Pomegranate", 3],
    ["Pear", 3],
    ["Tangerine", 3],
    ["Blueberry", 3],
    ["Mango", 3],
    ["Watermelon", 3],
    ["Blackberry", 3],
    ["Orange", 3],
    ["Raspberry", 3],
    ["Banana", 3]
]

for r in data:
    ws4.append(r)

# ws4.auto_filter.ref = "A1:B10"
ws4.auto_filter.ref = "A1:{}{}".format(get_column_letter(ws4.max_column),
                                       ws4.max_row)
# print(ws4.auto_filter.ref)
# ws4.auto_filter.add_filter_column(0, " ")
# ws4.auto_filter.add_sort_condition("B2:B15")

print(ws4.max_column)

print(get_column_letter(ws4.max_column))

try:
    wb.save(filename=dest_filename + ".xlsx")
except IOError:
    wb.save(filename=dest_filename + "_2.xlsx")
