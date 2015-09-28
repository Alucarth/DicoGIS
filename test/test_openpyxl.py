# -*- coding: UTF-8 -*-
#!/usr/bin/env python
from __future__ import unicode_literals
# -----------------------------------------------------------------------------
# Name:         DicoGIS
# Purpose:      Automatize the creation of a dictionnary of geographic data
#                   contained in a folders structures.
#                   It produces an Excel output file (.xls)
#
# Author:       Julien Moura (https://twitter.com/geojulien)
#
# Python:       2.7.x
# Created:      14/02/2013
# Updated:      13/04/2015
#
# Licence:      GPL 3
# ------------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Basic configuration
# self.book = Workbook(encoding='utf8')
book = Workbook()
# self.book.set_owner(str('DicoGIS_') + str(DGversion))


# Some customization: fonts and styles
# headers style

header_ft = Font(bold=True)
header_al = Alignment(horizontal='right', vertical='center')


# # hyperlinks style
# self.url = easyxf(u'font: underline single')
# # errors style
# self.xls_erreur = easyxf('pattern: pattern solid, fore_colour red;'
#                          'font: colour white, bold True;')
# # cell style handling return in-cell
# self.xls_wrap = easyxf('align: wrap True')
# # date cell style
# self.xls_date = easyxf(num_format_str='DD/MM/YYYY')



#     """ adding a new sheet for metrics """
# sheet
# ws = book.create_sheet()
ws = book.active
ws.title = "Metrics"


# properties
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
ws.page_setup.fitToHeight = 0
ws.page_setup.fitToWidth = 1
ws.print_options.horizontalCentered = True
ws.print_options.verticalCentered = True
ws.header_footer.center_header.text = "Global metrics"


row_count = ws.get_highest_row() - 1
print(row_count)

# first line: headers
row = ws.row_dimensions[1]
row.font = header_ft
row.alignment = header_al

# headers
ws['A1'] = "Youhou"


# self.feuySTATS.write(0, 0, "Totals", self.entete)
# self.feuySTATS.write(0, 1, "=== Global Statistics ===", self.entete)
# self.feuySTATS.write(1, 0, self.blabla.get('feats_class'), self.entete)
# self.feuySTATS.write(2, 0, self.blabla.get('num_attrib'), self.entete)
# self.feuySTATS.write(3, 0, self.blabla.get('num_objets'), self.entete)
# self.feuySTATS.write(4, 0, self.blabla.get('gdal_warn'), self.entete)
# self.feuySTATS.write(6, 0, self.blabla.get('geometrie'), self.entete)
# self.logger.info('Sheet for global statistics adedd')
# # tunning headers
# # lg_shp_names = [len(lg) for lg in self.li_shp]
# # lg_tab_names = [len(lg) for lg in self.li_tab]
# # self.feuySTATS.col(0).width = max(lg_shp_names + lg_tab_names) * 100
# # self.feuySTATS.col(1).width = len(self.blabla.get('browse')) * 256
# # self.feuySTATS.col(9).width = 35 * 256


# # """ adding a new sheet for vectors informations """
# # sheet
# self.feuyVC = self.book.add_sheet(self.blabla.get('sheet_vectors'),
#                                   cell_overwrite_ok=True)
# # headers
# self.feuyVC.write(0, 0, self.blabla.get('nomfic'), self.entete)
# self.feuyVC.write(0, 1, self.blabla.get('path'), self.entete)
# self.feuyVC.write(0, 2, self.blabla.get('theme'), self.entete)
# self.feuyVC.write(0, 3, self.blabla.get('num_attrib'), self.entete)
# self.feuyVC.write(0, 4, self.blabla.get('num_objets'), self.entete)
# self.feuyVC.write(0, 5, self.blabla.get('geometrie'), self.entete)
# self.feuyVC.write(0, 6, self.blabla.get('srs'), self.entete)
# self.feuyVC.write(0, 7, self.blabla.get('srs_type'), self.entete)
# self.feuyVC.write(0, 8, self.blabla.get('codepsg'), self.entete)
# self.feuyVC.write(0, 9, self.blabla.get('emprise'), self.entete)
# self.feuyVC.write(0, 10, self.blabla.get('date_crea'), self.entete)
# self.feuyVC.write(0, 11, self.blabla.get('date_actu'), self.entete)
# self.feuyVC.write(0, 12, self.blabla.get('format'), self.entete)
# self.feuyVC.write(0, 13, self.blabla.get('li_depends'), self.entete)
# self.feuyVC.write(0, 14, self.blabla.get('tot_size'), self.entete)
# self.feuyVC.write(0, 15, self.blabla.get('li_chps'), self.entete)
# self.feuyVC.write(0, 16, self.blabla.get('gdal_warn'), self.entete)
# self.logger.info('Sheet vectors adedd')
# # tunning headers
# lg_shp_names = [len(lg) for lg in self.li_shp]
# lg_tab_names = [len(lg) for lg in self.li_tab]
# self.feuyVC.col(0).width = max(lg_shp_names + lg_tab_names) * 100
# self.feuyVC.col(1).width = len(self.blabla.get('browse')) * 256
# self.feuyVC.col(9).width = 35 * 256
# # freezing headers line and first column
# self.feuyVC.set_panes_frozen(True)
# self.feuyVC.set_horz_split_pos(1)
# self.feuyVC.set_vert_split_pos(1)

c = ws['A2']
ws.freeze_panes = c

book.save('test_openpyxl.xlsx')
