# -*- coding: UTF-8 -*-
#!/usr/bin/env python
from __future__ import (absolute_import, print_function, unicode_literals)
# ------------------------------------------------------------------------------
# Name:         Isogeo to Microsoft Excel 2010
# Purpose:      Get metadatas from an Isogeo share and store it into
#               a Excel worksheet. It's one of the submodules of
#               isogeo2office (https://bitbucket.org/isogeo/isogeo-2-office).
#
# Author:       Julien Moura (@geojulien)
#
# Python:       2.7.x
# Created:      14/08/2014
# Updated:      28/10/2016
# ------------------------------------------------------------------------------

# ##############################################################################
# ########## Libraries #############
# ##################################

# Standard library
import logging
from os import path

# Python 3 backported
from collections import OrderedDict     # ordered dictionary

# 3rd party library
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties

# ##############################################################################
# ############ Globals ############
# #################################

# LOG
logger = logging.getLogger("DicoGIS")

# ##############################################################################
# ########## Classes ###############
# ##################################


class md2xlsx(Workbook):
    """Export into a XLSX worksheet."""
    li_cols_vector = ["nomfic",
                      "path",
                      "theme",
                      "num_attrib",
                      "num_objets",
                      "geometrie",
                      "srs",
                      "srs_type",
                      "codepsg",
                      "emprise",
                      "date_crea",
                      "date_actu",
                      "format",
                      "li_depends",
                      "tot_size",
                      "li_chps",
                      "gdal_warn"
                      ]

    li_cols_raster = ["nomfic",
                      "path",
                      "theme",
                      "size_Y",
                      "size_X",
                      "pixel_w",
                      "pixel_h",
                      "origin_x",
                      "origin_y",
                      "srs_type",
                      "codepsg",
                      "emprise",
                      "date_crea",
                      "date_actu",
                      "num_bands",
                      "format",
                      "compression",
                      "coloref",
                      "li_depends",
                      "tot_size",
                      "gdal_warn",
                      ]

    li_cols_filedb = ["nomfic",
                      "path",
                      "theme",
                      "tot_size",
                      "date_crea",
                      "date_actu",
                      "feats_class",
                      "num_attrib",
                      "num_objets",
                      "geometrie",
                      "srs",
                      "srs_type",
                      "codepsg",
                      "emprise",
                      "li_chps",
                      ]

    li_cols_mapdocs = ["nomfic",
                       "path",
                       "theme",
                       "custom_title",
                       "creator_prod",
                       "keywords",
                       "subject",
                       "res_image",
                       "tot_size",
                       "date_crea",
                       "date_actu",
                       "origin_x",
                       "origin_y",
                       "srs",
                       "srs_type",
                       "codepsg",
                       "sub_layers",
                       "num_attrib",
                       "num_objets",
                       "li_chps",
                       ]

    li_cols_lyr = ["nomfic",
                   "path",
                   "theme",
                   "custom_title",
                   "lyr_type",
                   "description",
                   "keywords",
                   "subject",
                   "license",
                   "tot_size",
                   "date_crea",
                   "date_actu",
                   "emprise",
                   "srs",
                   "srs_type",
                   "codepsg",
                   "sub_layers",
                   "num_attrib",
                   "num_objets",
                   "li_chps",
                   ]

    li_cols_caodao = ["nomfic",
                      "path",
                      "theme",
                      "tot_size",
                      "date_crea",
                      "date_actu",
                      "sub_layers",
                      "num_attrib",
                      "num_objets",
                      "geometrie",
                      "srs",
                      "srs_type",
                      "codepsg",
                      "emprise",
                      "li_chps",
                      ]

    li_cols_sgbd = ["nomfic",
                    "conn_chain",
                    "schema",
                    "num_attrib",
                    "num_objets",
                    "geometrie",
                    "srs",
                    "srs_type",
                    "codepsg",
                    "emprise",
                    "format",
                    "li_chps",
                    "gdal_err"
                    ]

    def __init__(self, lang="EN", texts=OrderedDict()):
        """TO DOC.

        Keyword arguments:
        """
        super(md2xlsx, self).__init__()
        # super(files2xlsx, self).__init__(write_only=True)
        self.txt = texts

        # styles
        s_date = NamedStyle(name="date")
        s_date.number_format = "dd/mm/yyyy"
        s_wrap = NamedStyle(name="wrap")
        s_wrap.alignment = Alignment(wrap_text=True)
        self.add_named_style(s_date)
        self.add_named_style(s_wrap)

        # deleting the default worksheet
        ws = self.active
        self.remove_sheet(ws)

    # ------------ Setting workbook ---------------------

    def set_worksheets(self, has_vector=0, has_raster=0, has_filedb=0,
                       has_mapdocs=0, has_cad=0, has_lyr=0, has_sgbd=0):
        """Add news sheets depending on present metadata types."""
        # SHEETS & HEADERS
        if has_vector and self.txt.get("sheet_vectors") not in self.sheetnames:
            self.ws_v = self.create_sheet(title=self.txt.get("sheet_vectors"))
            # headers
            self.ws_v.append([self.txt.get(i) for i in self.li_cols_vector])
            # styling
            for i in self.li_cols_vector:
                self.ws_v.cell(row=1,
                               column=self.li_cols_vector.index(i) + 1).style = "Headline 2"

            # initialize line counter
            self.idx_v = 1
        else:
            pass

        if has_raster and self.txt.get("sheet_rasters") not in self.sheetnames:
            self.ws_r = self.create_sheet(title=self.txt.get("sheet_rasters"))
            # headers
            self.ws_r.append([self.txt.get(i) for i in self.li_cols_raster])
            # styling
            for i in self.li_cols_raster:
                self.ws_r.cell(row=1,
                               column=self.li_cols_raster.index(i) + 1).style = "Headline 2"

            # initialize line counter
            self.idx_r = 1
        else:
            pass

        if has_filedb and self.txt.get("sheet_filedb") not in self.sheetnames:
            self.ws_fdb = self.create_sheet(title=self.txt.get("sheet_filedb"))
            # headers
            self.ws_fdb.append([self.txt.get(i) for i in self.li_cols_filedb])
            # styling
            for i in self.li_cols_filedb:
                self.ws_fdb.cell(row=1,
                                 column=self.li_cols_filedb.index(i) + 1).style = "Headline 2"

            # initialize line counter
            self.idx_f = 1
        else:
            pass

        if has_mapdocs and self.txt.get("sheet_maplans") not in self.sheetnames:
            self.ws_mdocs = self.create_sheet(title=self.txt.get("sheet_maplans"))
            # headers
            self.ws_mdocs.append([self.txt.get(i) for i in self.li_cols_mapdocs])
            # styling
            for i in self.li_cols_mapdocs:
                self.ws_mdocs.cell(row=1,
                                   column=self.li_cols_mapdocs.index(i) + 1).style = "Headline 2"

            # initialize line counter
            self.idx_m = 1
        else:
            pass

        if has_cad and self.txt.get("sheet_cdao") not in self.sheetnames:
            self.ws_cad = self.create_sheet(title=self.txt.get("sheet_cdao"))
            # headers
            self.ws_cad.append([self.txt.get(i) for i in self.li_cols_caodao])
            # styling
            for i in self.li_cols_caodao:
                self.ws_cad.cell(row=1,
                                 column=self.li_cols_caodao.index(i) + 1).style = "Headline 2"

            # initialize line counter
            self.idx_c = 1
        else:
            pass

        if has_sgbd and "PostGIS" not in self.sheetnames:
            self.ws_sgbd = self.create_sheet(title="PostGIS")
            # headers
            self.ws_sgbd.append([self.txt.get(i) for i in self.li_cols_sgbd])
            # styling
            for i in self.li_cols_sgbd:
                self.ws_sgbd.cell(row=1,
                                  column=self.li_cols_sgbd.index(i) + 1).style = "Headline 2"
            # initialize line counter
            self.idx_s = 1
        else:
            pass

        if has_lyr and "LYR" not in self.sheetnames:
            self.ws_lyr = self.create_sheet(title="PostGIS")
            # headers
            self.ws_lyr.append([self.txt.get(i) for i in self.li_cols_lyr])
            # styling
            for i in self.li_cols_lyr:
                self.ws_lyr.cell(row=1,
                                 column=self.li_cols_lyr.index(i) + 1).style = "Headline 2"
            # initialize line counter
            self.idx_l = 1
        else:
            pass

        # end of method
        return

    def tunning_worksheets(self):
        """Clean up and tunning worksheet."""
        for sheet in self.worksheets:
            # Freezing panes
            c_freezed = sheet['B2']
            sheet.freeze_panes = c_freezed

            # Print properties
            sheet.print_options.horizontalCentered = True
            sheet.print_options.verticalCentered = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

            # Others properties
            wsprops = sheet.sheet_properties
            wsprops.filterMode = True

            # enable filters
            sheet.auto_filter.ref = str("A1:{}{}").format(get_column_letter(sheet.max_column),
                                                          sheet.max_row)
            # columns width
            sheet.column_dimensions['A'].bestFit = True
            # sheet.column_dimensions['A'].auto_size = True
            # sheet.column_dimensions['B'].auto_size = True

            dims = {}
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        val = unicode(cell.value)
                        dims[cell.column] = max((dims.get(cell.column, 0), len(val)))
            for col, value in dims.items():
                sheet.column_dimensions[col].width = value
        # end of method
        return

    # ------------ Writing metadata ---------------------

    def store_md_vector(self, layer):
        """Store metadata about a vector dataset."""
        # increment line
        self.idx_v += 1
        # local variables
        champs = ""

        # in case of a source error
        if "error" in layer:
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.txt.get(layer.get('error'))
            logger.warning('\tproblem detected')
            self.ws_v["A{}".format(self.idx_v)] = layer.get('name')
            self.ws_v["A{}".format(self.idx_v)].style = "Warning Text"
            link = r'=HYPERLINK("{0}","{1}")'.format(layer.get(u'folder'),
                                                     self.txt.get('browse'))
            self.ws_v["B{}".format(self.idx_v)] = link
            self.ws_v["B{}".format(self.idx_v)].style = "Warning Text"
            self.ws_v["C{}".format(self.idx_v)] = err_mess
            self.ws_v["C{}".format(self.idx_v)].style = "Warning Text"
            # gdal info
            if "err_gdal" in layer:
                logger.warning('\tproblem detected')
                self.ws_v["Q{}".format(self.idx_v)] = "{0} : {1}".format(layer.get('err_gdal')[0],
                                                                         layer.get('err_gdal')[1])
                self.ws_v["Q{}".format(self.idx_v)].style = "Warning Text"
            else:
                pass
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_v["A{}".format(self.idx_v)] = layer.get('name')

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{0}","{1}")'.format(layer.get(u'folder'),
                                                 self.txt.get('browse'))
        self.ws_v["B{}".format(self.idx_v)] = link
        self.ws_v["B{}".format(self.idx_v)].style = "Hyperlink"

        # Name of parent folder with an exception if this is the format name
        self.ws_v["C{}".format(self.idx_v)] = path.basename(layer.get(u'folder'))

        # Fields count
        self.ws_v["D{}".format(self.idx_v)] = layer.get(u'num_fields', "")
        # Objects count
        self.ws_v["E{}".format(self.idx_v)] = layer.get(u'num_obj', "")
        # Geometry type
        self.ws_v["F{}".format(self.idx_v)] = layer.get(u'type_geom', "")
        # Name of srs
        self.ws_v["G{}".format(self.idx_v)] = layer.get(u'srs', "")
        # Type of SRS
        self.ws_v["H{}".format(self.idx_v)] = layer.get(u'srs_type', "")
        # EPSG code
        self.ws_v["I{}".format(self.idx_v)] = layer.get(u'EPSG', "")
        # Spatial extent
        emprise = "Xmin : {0} - Xmax : {1} | \nYmin : {2} - Ymax : {3}"\
                  .format(unicode(layer.get(u'Xmin')),
                          unicode(layer.get(u'Xmax')),
                          unicode(layer.get(u'Ymin')),
                          unicode(layer.get(u'Ymax')))
        self.ws_v["J{}".format(self.idx_v)].style = "wrap"
        self.ws_v["J{}".format(self.idx_v)] = emprise

        # Creation date
        self.ws_v["K{}".format(self.idx_v)] = layer.get(u'date_crea')
        # Last update date
        self.ws_v["L{}".format(self.idx_v)] = layer.get(u'date_actu')
        # Format of data
        self.ws_v["M{}".format(self.idx_v)] = layer.get(u'type')
        # dependencies
        self.ws_v["N{}".format(self.idx_v)].style = "wrap"
        self.ws_v.cell("N{}".format(self.idx_v)).value = u' |\n '.join(layer.get('dependencies', []))
        # total size
        self.ws_v["O{}".format(self.idx_v)] = layer.get(u'total_size')

        # Field informations
        fields = layer.get("fields")
        for chp in fields.keys():
            # field type
            if 'Integer' in fields[chp][0]:
                tipo = self.txt.get(u'entier')
            elif fields[chp][0] == 'Real':
                tipo = self.txt.get(u'reel')
            elif fields[chp][0] == 'String':
                tipo = self.txt.get(u'string')
            elif fields[chp][0] == 'Date':
                tipo = self.txt.get(u'date')
            else:
                tipo = "unknown"
                logger.warning(chp + " unknown type")

            # concatenation of field informations
            try:
                champs = champs + chp +\
                          u" (" + tipo + self.txt.get(u'longueur') +\
                          unicode(fields[chp][1]) +\
                          self.txt.get(u'precision') +\
                          unicode(fields[chp][2]) + u") ; "
            except UnicodeDecodeError:
                logger.warning('Field name with special letters: {}'.format(chp.decode('latin1')))
                # decode the fucking field name
                champs = champs + chp.decode('latin1') \
                + u" ({}, Lg. = {}, Pr. = {}) ;".format(tipo,
                                                        fields[chp][1],
                                                        fields[chp][2])
                # then continue
                continue

        # Once all fieds explored, write them
        self.ws_v["P{}".format(self.idx_v)] = champs

        # end of method
        return

    def store_md_raster(self, layer, bands):
        """Store metadata about a raster dataset."""
        # increment line
        self.idx_r += 1

        # in case of a source error
        if "error" in layer:
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.txt.get(layer.get('error'))
            logger.warning('\tproblem detected')
            self.ws_r["A{}".format(self.idx_r)] = layer.get('name')
            link = r'=HYPERLINK("{0}","{1}")'.format(layer.get(u'folder'),
                                                     self.txt.get('browse'))
            self.ws_r["B{}".format(self.idx_r)] = link
            self.ws_r["B{}".format(self.idx_r)].style = "Warning Text"
            self.ws_r["C{}".format(self.idx_r)] = err_mess
            self.ws_r["C{}".format(self.idx_r)].style = "Warning Text"
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_r["A{}".format(self.idx_r)] = layer.get('name')

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{0}","{1}")'.format(layer.get(u'folder'),
                                                 self.txt.get('browse'))
        self.ws_r["B{}".format(self.idx_r)] = link
        self.ws_r["B{}".format(self.idx_r)].style = "Hyperlink"

        # Name of parent folder with an exception if this is the format name
        self.ws_r["C{}".format(self.idx_r)] = path.basename(layer.get(u'folder'))

        # Image dimensions
        self.ws_r["D{}".format(self.idx_r)] = layer.get(u'num_rows')
        self.ws_r["E{}".format(self.idx_r)] = layer.get(u'num_cols')

        # Pixel dimensions
        self.ws_r["F{}".format(self.idx_r)] = layer.get(u'pixelWidth')
        self.ws_r["G{}".format(self.idx_r)] = layer.get(u'pixelHeight')

        # Image dimensions
        self.ws_r["H{}".format(self.idx_r)] = layer.get(u'xOrigin')
        self.ws_r["I{}".format(self.idx_r)] = layer.get(u'yOrigin')

        # Type of SRS
        self.ws_r["J{}".format(self.idx_r)] = layer.get(u'srs_type')
        # EPSG code
        self.ws_r["K{}".format(self.idx_r)] = layer.get(u'EPSG')

        # Creation date
        self.ws_r["M{}".format(self.idx_r)] = layer.get(u'date_crea')
        # Last update date
        self.ws_r["N{}".format(self.idx_r)] = layer.get(u'date_actu')

        # Number of bands
        self.ws_r["O{}".format(self.idx_r)] = layer.get(u'num_bands')

        # Format of data
        self.ws_r["P{}".format(self.idx_r)] = "{0} {1}".format(layer.get(u'format'),
                                                               layer.get('format_version'))
        # Compression rate
        self.ws_r["Q{}".format(self.idx_r)] = layer.get(u'compr_rate')

        # Color referential
        self.ws_r["R{}".format(self.idx_r)] = layer.get(u'color_ref')

        # Dependencies
        self.ws_r["S{}".format(self.idx_v)].style = "wrap"
        self.ws_r.cell("S{}".format(self.idx_v)).value = u' |\n '.join(layer.get(u'dependencies'))

        # total size of file and its dependencies
        self.ws_r["T{}".format(self.idx_r)] = layer.get(u'total_size')

        # in case of a source error
        if layer.get('err_gdal', [0, ])[0] != 0:
            logger.warning('\tproblem detected')
            self.ws_r["U{}".format(self.idx_r)] = "{0} : {1}".format(layer.get('err_gdal')[0],
                                                                     layer.get('err_gdal')[1])
            self.ws_r["U{}".format(self.idx_r)].style = "Warning Text"
        else:
            pass

        # end of method
        return

    def store_md_fdb(self, filedb):
        """Storing metadata about a file database."""
        # increment line
        self.idx_f += 1

        # in case of a source error
        if "error" in filedb:
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.txt.get(filedb.get('error'))
            logger.warning('\tproblem detected')
            self.ws_fdb["A{}".format(self.idx_f)] = filedb.get('name')
            link = r'=HYPERLINK("{0}","{1}")'.format(filedb.get(u'folder'),
                                                     self.txt.get('browse'))
            self.ws_fdb["B{}".format(self.idx_f)] = link
            self.ws_fdb["B{}".format(self.idx_f)].style = "Warning Text"
            self.ws_fdb["C{}".format(self.idx_f)] = err_mess
            self.ws_fdb["C{}".format(self.idx_f)].style = "Warning Text"
            # gdal info
            if "err_gdal" in filedb:
                logger.warning('\tproblem detected')
                self.ws_fdb["Q{}".format(self.idx_v)] = "{0} : {1}"\
                                                        .format(filedb.get('err_gdal')[0],
                                                                filedb.get('err_gdal')[1])
                self.ws_fdb["Q{}".format(self.idx_v)].style = "Warning Text"
            else:
                pass
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_fdb["A{}".format(self.idx_f)] = filedb.get('name')

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{0}","{1}")'.format(filedb.get(u'folder'),
                                                 self.txt.get('browse'))
        self.ws_fdb["B{}".format(self.idx_f)] = link
        self.ws_fdb["B{}".format(self.idx_f)].style = "Hyperlink"

        self.ws_fdb["C{}".format(self.idx_f)] = path.basename(filedb.get(u'folder'))
        self.ws_fdb["D{}".format(self.idx_f)] = filedb.get(u'total_size')
        self.ws_fdb["E{}".format(self.idx_f)] = filedb.get(u'date_crea')
        self.ws_fdb["F{}".format(self.idx_f)] = filedb.get(u'date_actu')
        self.ws_fdb["G{}".format(self.idx_f)] = filedb.get(u'layers_count')
        self.ws_fdb["H{}".format(self.idx_f)] = filedb.get(u'total_fields')
        self.ws_fdb["I{}".format(self.idx_f)] = filedb.get(u'total_objs')

        # parsing layers
        for (layer_idx, layer_name) in zip(filedb.get(u'layers_idx'),
                                           filedb.get(u'layers_names')):
            # increment line
            self.idx_f += 1
            champs = ""
            # get the layer informations
            try:
                gdb_layer = filedb.get('{0}_{1}'.format(layer_idx,
                                                        layer_name))
            except UnicodeDecodeError:
                gdb_layer = filedb.get('{0}_{1}'.format(layer_idx,
                                                        unicode(layer_name.decode('latin1'))))
            # in case of a source error
            if gdb_layer.get('error'):
                err_mess = self.txt.get(gdb_layer.get('error'))
                logger.warning('\tproblem detected: \
                                  {0} in {1}'.format(err_mess,
                                                     gdb_layer.get(u'title')))
                self.ws_fdb["G{}".format(self.idx_f)] = gdb_layer.get(u'title')
                self.ws_fdb["G{}".format(self.idx_f)].style = "Warning Text"
                self.ws_fdb["H{}".format(self.idx_f)] = err_mess
                self.ws_fdb["H{}".format(self.idx_f)].style = "Warning Text"
                # Interruption of function
                continue
            else:
                pass

            self.ws_fdb["G{}".format(self.idx_f)] = gdb_layer.get('title')
            self.ws_fdb["H{}".format(self.idx_f)] = gdb_layer.get(u'num_fields')
            self.ws_fdb["I{}".format(self.idx_f)] = gdb_layer.get(u'num_obj')
            self.ws_fdb["J{}".format(self.idx_f)] = gdb_layer.get(u'type_geom')
            self.ws_fdb["K{}".format(self.idx_f)] = gdb_layer.get(u'srs')
            self.ws_fdb["L{}".format(self.idx_f)] = gdb_layer.get(u'srs_type')
            self.ws_fdb["M{}".format(self.idx_f)] = gdb_layer.get(u'EPSG')

            # Spatial extent
            emprise = u"Xmin : {0} - Xmax : {1} | \nYmin : {2} - Ymax : {3}"\
                      .format(unicode(gdb_layer.get(u'Xmin')),
                              unicode(gdb_layer.get(u'Xmax')),
                              unicode(gdb_layer.get(u'Ymin')),
                              unicode(gdb_layer.get(u'Ymax')))
            self.ws_fdb["N{}".format(self.idx_f)].style = "wrap"
            self.ws_fdb["N{}".format(self.idx_f)] = emprise

            # Field informations
            fields = gdb_layer.get(u'fields')
            for chp in fields.keys():
                # field type
                if 'Integer' in fields[chp][0]:
                    tipo = self.txt.get(u'entier')
                elif fields[chp][0] == 'Real':
                    tipo = self.txt.get(u'reel')
                elif fields[chp][0] == 'String':
                    tipo = self.txt.get(u'string')
                elif fields[chp][0] == 'Date':
                    tipo = self.txt.get(u'date')
                else:
                    tipo = "unknown"
                    logger.warning(chp + " unknown type")

                # concatenation of field informations
                try:
                    champs = champs + chp +\
                             u" (" + tipo + self.txt.get(u'longueur') +\
                             unicode(fields[chp][1]) +\
                             self.txt.get(u'precision') +\
                             unicode(fields[chp][2]) + u") ; "
                except UnicodeDecodeError:
                    logger.warning('Field name with special letters: {}'
                                    .format(chp.decode('latin1')))
                    # decode the fucking field name
                    champs = champs + chp.decode('latin1') \
                    + u" ({}, Lg. = {}, Pr. = {}) ;".format(tipo,
                                                            fields[chp][1],
                                                            fields[chp][2])
                    # then continue
                    continue

            # Once all fieds explored, write them
            self.ws_fdb["O{}".format(self.idx_f)] = champs

        # end of method
        return

    def store_md_mapdoc(self, mapdoc):
        """To store mapdocs information from DicoGIS."""
        # increment line
        self.idx_m += 1

        # local variables
        champs = ""

        # in case of a source error
        if "error" in mapdoc:
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.txt.get(mapdoc.get('error'))
            logger.warning('\tproblem detected')
            self.ws_mdocs["A{}".format(self.idx_m)] = mapdoc.get('name')
            self.ws_mdocs["A{}".format(self.idx_m)].style = "Warning Text"
            link = r'=HYPERLINK("{0}","{1}")'.format(mapdoc.get(u'folder'),
                                                     self.txt.get('browse'))
            self.ws_mdocs["B{}".format(self.idx_m)] = link
            self.ws_mdocs["B{}".format(self.idx_m)].style = "Warning Text"
            self.ws_mdocs["C{}".format(self.idx_m)] = err_mess
            self.ws_mdocs["C{}".format(self.idx_m)].style = "Warning Text"
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_mdocs["A{}".format(self.idx_m)] = mapdoc.get('name')

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{0}","{1}")'.format(mapdoc.get(u'folder'),
                                                 self.txt.get('browse'))
        self.ws_mdocs["B{}".format(self.idx_m)] = link
        self.ws_mdocs["B{}".format(self.idx_m)].style = "Hyperlink"
        self.ws_mdocs["C{}".format(self.idx_m)] = path.dirname(mapdoc.get('folder'))
        self.ws_mdocs["D{}".format(self.idx_m)] = mapdoc.get('title')
        self.ws_mdocs["E{}".format(self.idx_m)] = mapdoc.get('creator_prod')
        self.ws_mdocs["F{}".format(self.idx_m)] = mapdoc.get('keywords')
        self.ws_mdocs["G{}".format(self.idx_m)] = mapdoc.get('subject')
        self.ws_mdocs["H{}".format(self.idx_m)] = mapdoc.get('dpi')
        self.ws_mdocs["I{}".format(self.idx_m)] = mapdoc.get('total_size')
        self.ws_mdocs["J{}".format(self.idx_m)] = mapdoc.get('date_crea')
        self.ws_mdocs["K{}".format(self.idx_m)] = mapdoc.get('date_actu')
        self.ws_mdocs["L{}".format(self.idx_m)] = mapdoc.get('xOrigin')
        self.ws_mdocs["M{}".format(self.idx_m)] = mapdoc.get('yOrigin')
        self.ws_mdocs["N{}".format(self.idx_m)] = mapdoc.get('srs')
        self.ws_mdocs["O{}".format(self.idx_m)] = mapdoc.get('srs_type')
        self.ws_mdocs["P{}".format(self.idx_m)] = mapdoc.get('EPSG')
        self.ws_mdocs["Q{}".format(self.idx_m)] = mapdoc.get('layers_count')
        self.ws_mdocs["R{}".format(self.idx_m)] = mapdoc.get('total_fields')
        self.ws_mdocs["S{}".format(self.idx_m)] = mapdoc.get('total_objs')

        for (layer_idx, layer_name) in zip(mapdoc.get(u'layers_idx'),
                                           mapdoc.get(u'layers_names')):
            # increment line
            self.idx_m += 1
            champs = ""

            # get the layer informations
            try:
                mdoc_layer = mapdoc.get('{0}_{1}'.format(layer_idx,
                                                         layer_name))
            except UnicodeDecodeError:
                mdoc_layer = mapdoc.get('{0}_{1}'.format(layer_idx,
                                                         unicode(layer_name.decode('latin1'))))
            # in case of a source error
            if mdoc_layer.get('error'):
                err_mess = self.txt.get(mdoc_layer.get('error'))
                logger.warning('\tproblem detected: \
                                  {0} in {1}'.format(err_mess,
                                                     mdoc_layer.get(u'title')))
                self.ws_mdocs["Q{}".format(self.idx_f)] = mdoc_layer.get(u'title')
                self.ws_mdocs["Q{}".format(self.idx_f)].style = "Warning Text"
                self.ws_mdocs["R{}".format(self.idx_f)] = err_mess
                self.ws_mdocs["R{}".format(self.idx_f)].style = "Warning Text"
                # loop must go on
                continue
            else:
                pass
            # layer info
            self.ws_mdocs["Q{}".format(self.idx_m)] = mdoc_layer.get('title')
            self.ws_mdocs["R{}".format(self.idx_m)] = mdoc_layer.get('num_fields')
            self.ws_mdocs["S{}".format(self.idx_m)] = mdoc_layer.get('num_objs')

            # Field informations
            fields = mdoc_layer.get(u'fields')
            for chp in fields.keys():
                # field type
                if 'Integer' in fields[chp][0]:
                    tipo = self.txt.get(u'entier')
                elif fields[chp][0] == 'Real':
                    tipo = self.txt.get(u'reel')
                elif fields[chp][0] == 'String':
                    tipo = self.txt.get(u'string')
                elif fields[chp][0] == 'Date':
                    tipo = self.txt.get(u'date')
                else:
                    tipo = "unknown"
                    logger.warning(chp + " unknown type")

                # concatenation of field informations
                try:
                    champs = champs + chp +\
                              u" (" + tipo + self.txt.get(u'longueur') +\
                              unicode(fields[chp][1]) +\
                              self.txt.get(u'precision') +\
                              unicode(fields[chp][2]) + u") ; "
                except UnicodeDecodeError:
                    logger.warning('Field name with special letters: {}'.format(chp.decode('latin1')))
                    # decode the fucking field name
                    champs = champs + chp.decode('latin1') \
                    + u" ({}, Lg. = {}, Pr. = {}) ;".format(tipo,
                                                            fields[chp][1],
                                                            fields[chp][2])
                    # then continue
                    continue

            # Once all fieds explored, write them
            self.ws_fdb["T{}".format(self.idx_f)] = champs

        # end of method
        return

    def store_md_cad(self, cad):
        """Store metadata about CAD dataset."""
        # increment line
        self.idx_c += 1

        # local variables
        champs = ""

        # in case of a source error
        if "error" in cad:
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.txt.get(cad.get('error'))
            logger.warning('\tproblem detected')
            self.ws_cad["A{}".format(self.idx_c)] = cad.get('name')
            self.ws_cad["A{}".format(self.idx_c)].style = "Warning Text"
            link = r'=HYPERLINK("{0}","{1}")'.format(cad.get(u'folder'),
                                                     self.txt.get('browse'))
            self.ws_cad["B{}".format(self.idx_c)] = link
            self.ws_cad["B{}".format(self.idx_c)].style = "Warning Text"
            self.ws_cad["C{}".format(self.idx_c)] = err_mess
            self.ws_cad["C{}".format(self.idx_c)].style = "Warning Text"
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_cad["A{}".format(self.idx_c)] = cad.get('name')

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{0}","{1}")'.format(cad.get(u'folder'),
                                                 self.txt.get('browse'))
        self.ws_cad["B{}".format(self.idx_c)] = link
        self.ws_cad["B{}".format(self.idx_c)].style = "Hyperlink"

        # Name of parent folder with an exception if this is the format name
        self.ws_cad["C{}".format(self.idx_c)] = path.basename(cad.get(u'folder'))
        # total size
        self.ws_cad["D{}".format(self.idx_c)] = cad.get(u'total_size')
        # Creation date
        self.ws_cad["E{}".format(self.idx_c)] = cad.get(u'date_crea')
        # Last update date
        self.ws_cad["F{}".format(self.idx_c)] = cad.get(u'date_actu')
        self.ws_cad["G{}".format(self.idx_c)] = cad.get(u'layers_count')
        self.ws_cad["H{}".format(self.idx_c)] = cad.get(u'total_fields')
        self.ws_cad["I{}".format(self.idx_c)] = cad.get(u'total_objs')

        # parsing layers
        for (layer_idx, layer_name) in zip(cad.get(u'layers_idx'),
                                           cad.get(u'layers_names')):
            # increment line
            self.idx_c += 1
            champs = ""
            # get the layer informations
            try:
                layer = cad.get('{0}_{1}'.format(layer_idx,
                                                 layer_name))
            except UnicodeDecodeError:
                layer = cad.get('{0}_{1}'.format(layer_idx,
                                                 unicode(layer_name.decode('latin1'))))
            # in case of a source error
            if layer.get('error'):
                err_mess = self.txt.get(layer.get('error'))
                logger.warning("\tproblem detected: "
                                "{0} in {1}".format(err_mess,
                                                    layer.get(u'title')))
                self.ws_cad["G{}".format(self.idx_c)] = layer.get(u'title')
                self.ws_cad["G{}".format(self.idx_c)].style = "Warning Text"
                self.ws_cad["H{}".format(self.idx_c)] = err_mess
                self.ws_cad["H{}".format(self.idx_c)].style = "Warning Text"
                # Interruption of function
                continue
            else:
                pass

            self.ws_cad["G{}".format(self.idx_c)] = layer.get('title')
            self.ws_cad["H{}".format(self.idx_c)] = layer.get(u'num_fields')
            self.ws_cad["I{}".format(self.idx_c)] = layer.get(u'num_obj')
            self.ws_cad["J{}".format(self.idx_c)] = layer.get(u'type_geom')
            self.ws_cad["K{}".format(self.idx_c)] = layer.get(u'srs')
            self.ws_cad["L{}".format(self.idx_c)] = layer.get(u'srs_type')
            self.ws_cad["M{}".format(self.idx_c)] = layer.get(u'EPSG')

            # Spatial extent
            emprise = u"Xmin : {0} - Xmax : {1} | \nYmin : {2} - Ymax : {3}"\
                      .format(unicode(layer.get(u'Xmin')),
                              unicode(layer.get(u'Xmax')),
                              unicode(layer.get(u'Ymin')),
                              unicode(layer.get(u'Ymax')))
            self.ws_cad["N{}".format(self.idx_c)].style = "wrap"
            self.ws_cad["N{}".format(self.idx_c)] = emprise

            # Field informations
            fields = layer.get(u'fields')
            for chp in fields.keys():
                # field type
                if 'Integer' in fields[chp][0]:
                    tipo = self.txt.get(u'entier')
                elif fields[chp][0] == 'Real':
                    tipo = self.txt.get(u'reel')
                elif fields[chp][0] == 'String':
                    tipo = self.txt.get(u'string')
                elif fields[chp][0] == 'Date':
                    tipo = self.txt.get(u'date')
                else:
                    tipo = "unknown"
                    logger.warning(chp + " unknown type")

                # concatenation of field informations
                try:
                    champs = champs + chp \
                                    + u" (" + tipo + self.txt.get(u'longueur')\
                                    + unicode(fields[chp][1])\
                                    + self.txt.get(u'precision')\
                                    + unicode(fields[chp][2]) + u") ; "
                except UnicodeDecodeError:
                    logger.warning("Field name with special letters: {}"
                                    .format(chp.decode('latin1')))
                    # decode the fucking field name
                    champs = champs + chp.decode('latin1') \
                                    + u" ({}, Lg. = {}, Pr. = {}) ;"\
                                        .format(tipo,
                                                fields[chp][1],
                                                fields[chp][2])
                    # then continue
                    continue

            # Once all fieds explored, write them
            self.ws_cad["O{}".format(self.idx_c)] = champs

        # End of method
        return

    def store_md_sgdb(self, layer):
        """Storing metadata about a file database."""
        # increment line
        self.idx_s += 1
        # local variable
        champs = ""

        # layer name
        self.ws_sgbd["A{}".format(self.idx_s)] = layer.get('name')

        # connection string
        self.ws_sgbd["B{}".format(self.idx_s)] = "{}@{}:{}-{}"\
                                                 .format(layer.get('user'),
                                                         layer.get('sgbd_host'),
                                                         layer.get('sgbd_port'),
                                                         layer.get('db_name'))
        self.ws_sgbd["B{}".format(self.idx_s)].style = "Hyperlink"
        # schema
        self.ws_sgbd["C{}".format(self.idx_s)] = layer.get('folder')

        # in case of a source error
        if "error" in layer:
            self.ws_sgbd["D{}".format(self.idx_s)] = layer.get("error")
            self.ws_sgbd["A{}".format(self.idx_s)].style = "Warning Text"
            self.ws_sgbd["B{}".format(self.idx_s)].style = "Warning Text"
            self.ws_sgbd["C{}".format(self.idx_s)].style = "Warning Text"
            self.ws_sgbd["D{}".format(self.idx_s)].style = "Warning Text"
            # gdal info
            if "err_gdal" in layer:
                self.ws_v["M{}".format(self.idx_v)] = "{0} : {1}"\
                                                      .format(layer.get('err_gdal')[0],
                                                              layer.get('err_gdal')[1])
                self.ws_v["M{}".format(self.idx_v)].style = "Warning Text"
            else:
                pass
            # interruption of function
            return False
        else:
            pass

        # structure
        self.ws_sgbd["D{}".format(self.idx_s)] = layer.get('num_fields')
        self.ws_sgbd["E{}".format(self.idx_s)] = layer.get('num_obj')
        self.ws_sgbd["F{}".format(self.idx_s)] = layer.get('type_geom')

        # SRS
        self.ws_sgbd["G{}".format(self.idx_s)] = layer.get(u'srs')
        self.ws_sgbd["H{}".format(self.idx_s)] = layer.get(u'srs_type')
        self.ws_sgbd["I{}".format(self.idx_s)] = layer.get(u'EPSG')

        # Spatial extent
        emprise = u"Xmin : {0} - Xmax : {1} | \nYmin : {2} - Ymax : {3}"\
                  .format(unicode(layer.get('Xmin')),
                          unicode(layer.get('Xmax')),
                          unicode(layer.get('Ymin')),
                          unicode(layer.get('Ymax')))
        self.ws_sgbd["J{}".format(self.idx_s)].style = "wrap"
        self.ws_sgbd["J{}".format(self.idx_s)] = emprise

        # type
        self.ws_sgbd["K{}".format(self.idx_s)] = layer.get(u'type')

        # Field informations
        fields = layer.get("fields")
        for chp in fields.keys():
            # field type
            if 'Integer' in fields[chp][0]:
                tipo = self.txt.get(u'entier')
            elif fields[chp][0] == 'Real':
                tipo = self.txt.get(u'reel')
            elif fields[chp][0] == 'String':
                tipo = self.txt.get(u'string')
            elif fields[chp][0] == 'Date':
                tipo = self.txt.get(u'date')
            else:
                tipo = "unknown"
                logger.warning(chp + " unknown type")

            # concatenation of field informations
            try:
                champs = champs + chp +\
                          u" (" + tipo + self.txt.get(u'longueur') +\
                          unicode(fields[chp][1]) +\
                          self.txt.get(u'precision') +\
                          unicode(fields[chp][2]) + u") ; "
            except UnicodeDecodeError:
                logger.warning('Field name with special letters: {}'.format(chp.decode('latin1')))
                # decode the fucking field name
                champs = champs + chp.decode('latin1') \
                + u" ({}, Lg. = {}, Pr. = {}) ;".format(tipo,
                                                        fields[chp][1],
                                                        fields[chp][2])
                # then continue
                continue

        # Once all fieds explored, write them
        self.ws_sgbd["L{}".format(self.idx_s)] = champs

        # end of method
        return

    def store_md_lyr(self, layer):
        """To store mapdocs information from DicoGIS."""
        # increment line
        self.idx_l += 1

        # local variables
        champs = ""

        # in case of a source error
        if "error" in mapdoc:
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.txt.get(mapdoc.get('error'))
            logger.warning('\tproblem detected')
            self.ws_mdocs["A{}".format(self.idx_l)] = mapdoc.get('name')
            self.ws_mdocs["A{}".format(self.idx_l)].style = "Warning Text"
            link = r'=HYPERLINK("{0}","{1}")'.format(mapdoc.get(u'folder'),
                                                     self.txt.get('browse'))
            self.ws_mdocs["B{}".format(self.idx_l)] = link
            self.ws_mdocs["B{}".format(self.idx_l)].style = "Warning Text"
            self.ws_mdocs["C{}".format(self.idx_l)] = err_mess
            self.ws_mdocs["C{}".format(self.idx_l)].style = "Warning Text"
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_mdocs["A{}".format(self.idx_l)] = mapdoc.get('name')

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{0}","{1}")'.format(mapdoc.get(u'folder'),
                                                 self.txt.get('browse'))
        self.ws_mdocs["B{}".format(self.idx_l)] = link
        self.ws_mdocs["B{}".format(self.idx_l)].style = "Hyperlink"
        self.ws_mdocs["C{}".format(self.idx_l)] = path.dirname(mapdoc.get('folder'))
        self.ws_mdocs["D{}".format(self.idx_l)] = mapdoc.get('title')
        self.ws_mdocs["E{}".format(self.idx_l)] = mapdoc.get('creator_prod')
        self.ws_mdocs["F{}".format(self.idx_l)] = mapdoc.get('keywords')
        self.ws_mdocs["G{}".format(self.idx_l)] = mapdoc.get('subject')
        self.ws_mdocs["H{}".format(self.idx_l)] = mapdoc.get('dpi')
        self.ws_mdocs["I{}".format(self.idx_l)] = mapdoc.get('total_size')
        self.ws_mdocs["J{}".format(self.idx_l)] = mapdoc.get('date_crea')
        self.ws_mdocs["K{}".format(self.idx_l)] = mapdoc.get('date_actu')
        self.ws_mdocs["L{}".format(self.idx_l)] = mapdoc.get('xOrigin')
        self.ws_mdocs["M{}".format(self.idx_l)] = mapdoc.get('yOrigin')
        self.ws_mdocs["N{}".format(self.idx_l)] = mapdoc.get('srs')
        self.ws_mdocs["O{}".format(self.idx_l)] = mapdoc.get('srs_type')
        self.ws_mdocs["P{}".format(self.idx_l)] = mapdoc.get('EPSG')
        self.ws_mdocs["Q{}".format(self.idx_l)] = mapdoc.get('layers_count')
        self.ws_mdocs["R{}".format(self.idx_l)] = mapdoc.get('total_fields')
        self.ws_mdocs["S{}".format(self.idx_l)] = mapdoc.get('total_objs')

        for (layer_idx, layer_name) in zip(mapdoc.get(u'layers_idx'),
                                           mapdoc.get(u'layers_names')):
            # increment line
            self.idx_l += 1
            champs = ""

            # get the layer informations
            try:
                mdoc_layer = mapdoc.get('{0}_{1}'.format(layer_idx,
                                                         layer_name))
            except UnicodeDecodeError:
                mdoc_layer = mapdoc.get('{0}_{1}'.format(layer_idx,
                                                         unicode(layer_name.decode('latin1'))))
            # in case of a source error
            if mdoc_layer.get('error'):
                err_mess = self.txt.get(mdoc_layer.get('error'))
                logger.warning('\tproblem detected: \
                                  {0} in {1}'.format(err_mess,
                                                     mdoc_layer.get(u'title')))
                self.ws_mdocs["Q{}".format(self.idx_f)] = mdoc_layer.get(u'title')
                self.ws_mdocs["Q{}".format(self.idx_f)].style = "Warning Text"
                self.ws_mdocs["R{}".format(self.idx_f)] = err_mess
                self.ws_mdocs["R{}".format(self.idx_f)].style = "Warning Text"
                # loop must go on
                continue
            else:
                pass
            # layer info
            self.ws_mdocs["Q{}".format(self.idx_l)] = mdoc_layer.get('title')
            self.ws_mdocs["R{}".format(self.idx_l)] = mdoc_layer.get('num_fields')
            self.ws_mdocs["S{}".format(self.idx_l)] = mdoc_layer.get('num_objs')

            # Field informations
            fields = mdoc_layer.get(u'fields')
            for chp in fields.keys():
                # field type
                if 'Integer' in fields[chp][0]:
                    tipo = self.txt.get(u'entier')
                elif fields[chp][0] == 'Real':
                    tipo = self.txt.get(u'reel')
                elif fields[chp][0] == 'String':
                    tipo = self.txt.get(u'string')
                elif fields[chp][0] == 'Date':
                    tipo = self.txt.get(u'date')
                else:
                    tipo = "unknown"
                    logger.warning(chp + " unknown type")

                # concatenation of field informations
                try:
                    champs = champs + chp +\
                              u" (" + tipo + self.txt.get(u'longueur') +\
                              unicode(fields[chp][1]) +\
                              self.txt.get(u'precision') +\
                              unicode(fields[chp][2]) + u") ; "
                except UnicodeDecodeError:
                    logger.warning('Field name with special letters: {}'.format(chp.decode('latin1')))
                    # decode the fucking field name
                    champs = champs + chp.decode('latin1') \
                    + u" ({}, Lg. = {}, Pr. = {}) ;".format(tipo,
                                                            fields[chp][1],
                                                            fields[chp][2])
                    # then continue
                    continue

            # Once all fieds explored, write them
            self.ws_fdb["T{}".format(self.idx_f)] = champs

        # end of method
        return

    # def dictionarize_lyr(self, mapdoc_infos, sheet, line):
    #     u""" write the infos of the map document into the Excel workbook """
    #     # in case of a source error
    #     if mapdoc_infos.get('error'):
    #         logger.warning('\tproblem detected')
    #         # source name
    #         sheet.write(line, 0, mapdoc_infos.get('name'))
    #         # link to parent folder
    #         link = 'HYPERLINK("{0}"; "{1}")'.format(mapdoc_infos.get(u'folder'),
    #                                                 self.blabla.get('browse'))
    #         sheet.write(line, 1, Formula(link), self.url)
    #         sheet.write(line, 2, self.blabla.get(mapdoc_infos.get('error')),
    #                     self.xls_erreur)
    #         # incrementing line
    #         mapdoc_infos['layers_count'] = 0
    #         # exiting function
    #         return sheet, line
    #     else:
    #         pass

    #     # PDF source name
    #     sheet.write(line, 0, mapdoc_infos.get('name'))

    #     # Path of parent folder formatted to be a hyperlink
    #     try:
    #         link = 'HYPERLINK("{0}"; "{1}")'.format(mapdoc_infos.get(u'folder'),
    #                                                 self.blabla.get('browse'))
    #     except UnicodeDecodeError:
    #         # write a notification into the log file
    #         logger.warning('Path name with special letters: {}'.format(mapdoc_infos.get(u'folder').decode('utf8')))
    #         # decode the fucking path name
    #         link = 'HYPERLINK("{0}"; "{1}")'.format(mapdoc_infos.get(u'folder').decode('utf8'),
    #                                                 self.blabla.get('browse'))

    #     sheet.write(line, 1, Formula(link), self.url)

    #     # Name of parent folder
    #     sheet.write(line, 2, path.basename(mapdoc_infos.get(u'folder')))

    #     # Document title
    #     sheet.write(line, 3, mapdoc_infos.get(u'title'))

    #     # Type of lyr
    #     sheet.write(line, 4, mapdoc_infos.get(u'type'))

    #     # Type of lyr
    #     sheet.write(line, 5, mapdoc_infos.get(u'license'))

    #     # subject
    #     sheet.write(line, 6, mapdoc_infos.get(u'description'))

    #     # total size
    #     sheet.write(line, 8, mapdoc_infos.get(u'total_size'))

    #     # Creation date
    #     sheet.write(line, 9, mapdoc_infos.get(u'date_crea'), self.xls_date)
    #     # Last update date
    #     sheet.write(line, 10, mapdoc_infos.get(u'date_actu'), self.xls_date)

    #     if mapdoc_infos.get(u'type') in ['Feature', 'Raster']:
    #         # Spatial extent
    #         emprise = u"Xmin : {0} - Xmax : {1} \
    #                    \nYmin : {2} - Ymax : {3}".format(unicode(mapdoc_infos.get(u'Xmin')),
    #                                                      unicode(mapdoc_infos.get(u'Xmax')),
    #                                                      unicode(mapdoc_infos.get(u'Ymin')),
    #                                                      unicode(mapdoc_infos.get(u'Ymax'))
    #                                                      )
    #         sheet.write(line, 11, emprise, self.xls_wrap)

    #         # SRS name
    #         sheet.write(line, 13, mapdoc_infos.get(u'srs'))
    #         # Type of SRS
    #         sheet.write(line, 14, mapdoc_infos.get(u'srs_type'))
    #         # EPSG code
    #         sheet.write(line, 15, mapdoc_infos.get(u'EsriSRS')[0])
    #     else:
    #         pass

    #     if mapdoc_infos.get(u'type') == u'Group':
    #         # Layers count
    #         sheet.write(line, 16, mapdoc_infos.get(u'layers_count'))
    #         # layer's name
    #         sheet.write(line + 1, 16, ' ; '.join(mapdoc_infos.get(u'layers_names')))
    #     else:
    #         pass

    #     if mapdoc_infos.get(u'type') == u'Feature':
    #         # number of fields
    #         sheet.write(line, 17, mapdoc_infos.get(u'num_fields'))

    #         # number of objects
    #         sheet.write(line, 18, mapdoc_infos.get(u'num_obj'))

    #         # definition query
    #         sheet.write(line, 7, mapdoc_infos.get(u'defquery'))

    #         # fields domain
    #         fields_info = mapdoc_infos.get(u'fields')
    #         champs = ""
    #         for chp in fields_info.keys():
    #             tipo = fields_info.get(chp)[0]
    #             # concatenation of field informations
    #             try:
    #                 champs = champs + chp +\
    #                           u" (" + tipo + self.blabla.get(u'longueur') +\
    #                           unicode(fields_info.get(chp)[1]) +\
    #                           self.blabla.get(u'precision') +\
    #                           unicode(fields_info.get(chp)[2]) + u") ; "
    #             except UnicodeDecodeError:
    #                 # write a notification into the log file
    #                 self.dico_err[layer_infos.get('name')] = self.blabla.get(u'err_encod')\
    #                                                     + chp.decode('latin1') \
    #                                                     + u"\n\n"
    #                 logger.warning('Field name with special letters: {}'.format(chp.decode('latin1')))
    #                 # decode the fucking field name
    #                 champs = champs + chp.decode('latin1') \
    #                 + u" ({}, Lg. = {}, Pr. = {}) ;".format(tipo,
    #                                                         fields_info.get(chp)[1],
    #                                                         fields_info.get(chp)[2])
    #                 # then continue
    #                 continue

    #         # Once all fieds explored, write them
    #         sheet.write(line, 19, champs)

    #         # write layer's name into the log
    #         # logger.info('\t -- {0} = OK'.format(mapdoc_layer.get(u'title')))

    #     else:
    #         pass

    #     # End of function
    #     return self.feuyMAPS, line
# ############################################################################
# ##### Stand alone program ########
# ##################################

if __name__ == '__main__':
    """ Standalone execution and tests
    """
    # ------------ Specific imports ---------------------

    # ------------ REAL START ----------------------------
    wb = md2xlsx()
    wb.set_worksheets(has_vector=1,
                      has_raster=1,
                      has_service=1,
                      has_resource=1,
                      has_cad=1,
                      has_mapdocs=1,
                      has_filedb=1,
                      has_sgbd=1)

    # saving the test file
    wb.save(r"DicoGIS_files2xlsx.xlsx")
