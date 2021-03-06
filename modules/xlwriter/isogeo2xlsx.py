# -*- coding: UTF-8 -*-
#!/usr/bin/env python
from __future__ import absolute_import, print_function, unicode_literals

# ------------------------------------------------------------------------------
# Name:         Isogeo to Microsoft Excel 2010
# Purpose:      Get metadatas from an Isogeo share and store it into
#               a Excel worksheet. It's one of the submodules of
#               isogeo2office (https://bitbucket.org/isogeo/isogeo-2-office).
#
# Author:       Julien Moura (@geojulien)
#
# Python:       2.7.x
# Created:      14/08/2014
# Updated:      15/04/2016
# ------------------------------------------------------------------------------

# ##############################################################################
# ########## Libraries #############
# ##################################

# Standard library
from datetime import datetime

try:
    from itertools import zip_longest
except ImportError:
    from itertools import izip_longest as zip_longest
import logging
from os import path
import re
from xml.sax.saxutils import escape  # '<' -> '&lt;'

# 3rd party library
import arrow
from isogeo_pysdk import Isogeo
from isogeo_pysdk import IsogeoTranslator
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Alignment
from openpyxl.worksheet.properties import WorksheetProperties

# ##############################################################################
# ########## Classes ###############
# ##################################


class Isogeo2xlsx(Workbook):
    """ Used to store Isogeo API results into an Excel worksheet (.xlsx)
    """

    cols_v = [
        "Titre",  # A
        "Nom",  # B
        "Résumé",  # C
        "Emplacement",  # D
        "Groupe de travail",  # E
        "Mots-clés",  # F
        "Thématique(s) INSPIRE",  # G
        "Conformité INSPIRE",  # H
        "Contexte de collecte",  # I
        "Méthode de collecte",  # J
        "Début de validité",  # K
        "Fin de validité",  # L
        "Fréquence de mise à jour",  # M
        "Commentaire",  # N
        "Création",  # O
        "# mises à jour",  # P
        "Dernière mise à jour",  # Q
        "Publication",  # R
        "Format (version - encodage)",  # S
        "SRS (EPSG)",  # T
        "Emprise",  # U
        "Géométrie",  # V
        "Résolution",  # W
        "Echelle",  # X
        "# Objets",  # Y
        "# Attributs",  # Z
        "Attributs (A-Z)",  # AA
        "Spécifications",  # AB
        "Cohérence topologique",  # AC
        "Conditions",  # AD
        "Limitations",  # AE
        "# Contacts",  # AF
        "Points de contact",  # AG
        "Autres contacts",  # AH
        "Téléchargeable",  # AI
        "Visualisable",  # AJ
        "Autres",  # AK
        "Editer",  # AL
        "Consulter",  # AM
        "MD - ID",  # AN
        "MD - Création",  # AO
        "MD - Modification",  # AP
        "MD - Langue",  # AQ
    ]

    cols_r = [
        "Titre",  # A
        "Nom",  # B
        "Résumé",  # C
        "Emplacement",  # D
        "Groupe de travail",  # E
        "Mots-clés",  # F
        "Thématique(s) INSPIRE",  # G
        "Conformité INSPIRE",  # H
        "Contexte de collecte",  # I
        "Méthode de collecte",  # J
        "Début de validité",  # K
        "Fin de validité",  # L
        "Fréquence de mise à jour",  # M
        "Commentaire",  # N
        "Création",  # O
        "# mises à jour",  # P
        "Dernière mise à jour",  # Q
        "Publication",  # R
        "Format (version - encodage)",  # S
        "SRS (EPSG)",  # T
        "Emprise",  # U
        "Résolution",  # V
        "Echelle",  # W
        "Spécifications",  # X
        "Cohérence topologique",  # Y
        "Conditions",  # Z
        "Limitations",  # AA
        "# Contacts",  # AB
        "Points de contact",  # AC
        "Autres contacts",  # AD
        "Téléchargeable",  # AE
        "Visualisable",  # AF
        "Autres",  # AG
        "Editer",  # AH
        "Consulter",  # AI
        "MD - ID",  # AJ
        "MD - Création",  # AK
        "MD - Modification",  # AL
        "MD - Langue",  # AM
    ]

    cols_s = [
        "Titre",  # A
        "Nom",  # B
        "Résumé",  # C
        "Emplacement",  # D
        "Groupe de travail",  # E
        "Mots-clés",  # F
        "Conformité INSPIRE",  # G
        "Création",  # H
        "# mises à jour",  # I
        "Dernière mise à jour",  # J
        "Publication",  # K
        "Format (version)",  # L
        "Emprise",  # M
        "Spécifications",  # N
        "Conditions",  # O
        "Limitations",  # P
        "# Contacts",  # Q
        "Points de contact",  # R
        "Autres contacts",  # S
        "Téléchargeable",  # T
        "Visualisable",  # U
        "Autres",  # V
        "Editer",  # W
        "Consulter",  # X
        "MD - ID",  # Y
        "MD - Création",  # Z
        "MD - Modification",  # AA
        "MD - Langue",  # AB
    ]

    cols_rz = [
        "Titre",  # A
        "Résumé",  # B
        "Emplacement",  # C
        "Groupe de travail",  # D
        "Mots-clés",  # E
        "Création",  # F
        "# mises à jour",  # G
        "Dernière mise à jour",  # H
        "Publication",  # I
        "Format (version)",  # J
        "Conditions",  # K
        "Limitations",  # L
        "# Contacts",  # M
        "Points de contact",  # N
        "Autres contacts",  # O
        "Téléchargeable",  # P
        "Visualisable",  # Q
        "Autres",  # R
        "Editer",  # S
        "Consulter",  # T
        "MD - ID",  # U
        "MD - Création",  # V
        "MD - Modification",  # W
        "MD - Langue",  # X
    ]

    def __init__(self, lang="FR", url_base=""):
        """Instanciating the output workbook."""
        super(Isogeo2xlsx, self).__init__()
        # super(Isogeo2xlsx, self).__init__(write_only=True)

        # OpenCatalog url
        self.url_base = url_base

        # styles
        s_date = NamedStyle(name="date")
        s_wrap = NamedStyle(name="wrap")
        s_wrap.alignment = Alignment(wrap_text=True)
        self.add_named_style(s_date)
        self.add_named_style(s_wrap)

        # deleting the default worksheet
        ws = self.active
        self.remove_sheet(ws)

        # LOCALE
        if lang.lower() == "fr":
            s_date.number_format = "dd/mm/yyyy"
            self.dates_fmt = "DD/MM/YYYY"
            self.locale_fmt = "fr_FR"
        else:
            s_date.number_format = "yyyy/mm/dd"
            self.dates_fmt = "YYYY/MM/DD"
            self.locale_fmt = "uk_UK"
        # TRANSLATIONS
        self.tr = IsogeoTranslator(lang).tr

    # ------------ Setting workbook ---------------------

    def set_worksheets(self, auto=None, vector=1, raster=1, service=1, resource=1):
        """Adds new sheets depending on present metadata types.

        auto: typically auto=search_results.get('tags').keys()
        """
        if type(auto) == list:
            logging.info("Automatic sheets creation based on tags")
            if "type:vector-dataset" in auto:
                vector = 1
            else:
                vector = 0
            if "type:raster-dataset" in auto:
                raster = 1
            else:
                raster = 0
                pass
            if "type:resource" in auto:
                resource = 1
            else:
                resource = 0
                pass
            if "type:service" in auto:
                service = 1
            else:
                service = 0
                pass
        else:
            pass

        # SHEETS & HEADERS
        if vector:
            self.ws_v = self.create_sheet(title="Vecteurs")
            # headers
            self.ws_v.append([i for i in self.cols_v])
            # styling
            for i in self.cols_v:
                self.ws_v.cell(
                    row=1, column=self.cols_v.index(i) + 1
                ).style = "Headline 2"
            # initialize line counter
            self.idx_v = 1
            # log
            logging.info("Vectors sheet added")
        else:
            pass

        if raster:
            self.ws_r = self.create_sheet(title="Raster")
            # headers
            self.ws_r.append([i for i in self.cols_r])
            # styling
            for i in self.cols_r:
                self.ws_r.cell(
                    row=1, column=self.cols_r.index(i) + 1
                ).style = "Headline 2"
            # initialize line counter
            self.idx_r = 1
            # log
            logging.info("Rasters sheet added")
        else:
            pass

        if service:
            self.ws_s = self.create_sheet(title="Services")
            # headers
            self.ws_s.append([i for i in self.cols_s])
            # styling
            for i in self.cols_s:
                self.ws_s.cell(
                    row=1, column=self.cols_s.index(i) + 1
                ).style = "Headline 2"
            # initialize line counter
            self.idx_s = 1
            # log
            logging.info("Services sheet added")
        else:
            pass

        if resource:
            self.ws_rz = self.create_sheet(title="Ressources")
            # headers
            self.ws_rz.append([i for i in self.cols_rz])
            # styling
            for i in self.cols_rz:
                self.ws_rz.cell(
                    row=1, column=self.cols_rz.index(i) + 1
                ).style = "Headline 2"
            # initialize line counter
            self.idx_rz = 1
            # log
            logging.info("Resources sheet added")
        else:
            pass

        # end of method
        return

    # ------------ Writing metadata ---------------------

    def store_metadatas(self, metadata):
        """ TO DOCUMENT
        """
        if metadata.get("type") == "vectorDataset":
            self.idx_v += 1
            self.store_md_vector(metadata, self.ws_v, self.idx_v)
            return
        elif metadata.get("type") == "rasterDataset":
            self.idx_r += 1
            self.store_md_raster(metadata, self.ws_r, self.idx_r)
            return
        elif metadata.get("type") == "service":
            self.idx_s += 1
            self.store_md_service(metadata, self.ws_s, self.idx_s)
            return
        elif metadata.get("type") == "resource":
            self.idx_rz += 1
            self.store_md_resource(metadata, self.ws_rz, self.idx_rz)
            return
        else:
            print("Type of metadata is not recognized/handled: " + metadata.get("type"))
            pass
        # end of method
        return

    def store_md_vector(self, md, ws, idx):
        """ TO DOCUMENT
        """
        # variables
        tags = md.get("tags")

        # IDENTIFICATION
        ws["A{}".format(idx)] = md.get("title", "")
        ws["B{}".format(idx)] = md.get("name", "")
        ws["C{}".format(idx)] = md.get("abstract", "")

        # path to source
        src_path = md.get("path", "")
        if path.isfile(src_path):
            link_path = r'=HYPERLINK("{0}","{1}")'.format(
                path.dirname(src_path), src_path
            )
            ws["D{}".format(idx)] = link_path
            logging.info("Path reachable")
        else:
            ws["D{}".format(idx)] = src_path
            logging.info("Path not recognized nor reachable")
            pass

        # owner
        ws["E{}".format(idx)] = next(v for k, v in tags.items() if "owner:" in k)

        # KEYWORDS & INSPIRE THEMES
        keywords = []
        inspire = []
        if "keywords" in md.keys():
            for k in md.get("keywords"):
                if k.get("_tag").startswith("keyword:is"):
                    keywords.append(k.get("text"))
                elif k.get("_tag").startswith("keyword:in"):
                    inspire.append(k.get("text"))
                else:
                    logging.info("Unknown keyword type: " + k.get("_tag"))
                    continue
            ws["F{}".format(idx)] = " ;\n".join(sorted(keywords))
            ws["G{}".format(idx)] = " ;\n".join(sorted(inspire))
        else:
            logging.info("Vector dataset without any keyword or INSPIRE theme")

        # conformity
        ws["H{}".format(idx)] = "conformity:inspire" in tags

        # HISTORY
        ws["I{}".format(idx)] = md.get("collectionContext", "")
        ws["J{}".format(idx)] = md.get("collectionMethod", "")

        # validity
        if md.get("validFrom"):
            valid_start = arrow.get(md.get("validFrom"))
            valid_start = "{0}".format(valid_start.format("DD/MM/YYYY", "fr_FR"))
        else:
            valid_start = ""
        ws["K{}".format(idx)] = valid_start

        if md.get("validTo"):
            valid_end = arrow.get(md.get("validTo"))
            valid_end = "{0}".format(valid_end.format("DD/MM/YYYY", "fr_FR"))
        else:
            valid_end = ""
        ws["L{}".format(idx)] = valid_end

        ws["M{}".format(idx)] = md.get("updateFrequency", "")
        ws["N{}".format(idx)] = md.get("validComment", "")

        # EVENTS
        # data creation date
        if md.get("created"):
            data_created = arrow.get(md.get("created"))
            data_created = "{0}".format(data_created.format("DD/MM/YYYY", "fr_FR"))
        else:
            data_created = ""
        ws["O{}".format(idx)] = data_created

        # events count
        ws["P{}".format(idx)] = len(md.get("events", ""))

        # data last update
        if md.get("modified"):
            data_updated = arrow.get(md.get("created"))
            data_updated = "{0}".format(data_updated.format("DD/MM/YYYY", "fr_FR"))
        else:
            data_updated = ""
        ws["Q{}".format(idx)] = data_updated

        # TECHNICAL
        # format
        if "format" in md.keys():
            format_lbl = next(v for k, v in tags.items() if "format:" in k)
        else:
            format_lbl = ""
        ws["S{}".format(idx)] = "{0} ({1} - {2})".format(
            format_lbl, md.get("formatVersion", "NR"), md.get("encoding", "NR")
        )

        # SRS
        srs = md.setdefault("coordinate-system", {"name": "NR", "code": "NR"})
        ws["T{}".format(idx)] = "{0} ({1})".format(
            srs.get("name", "NR"), srs.get("code", "NR")
        )

        # bounding box
        bbox = md.get("envelope", None)
        if bbox:
            coords = bbox.get("coordinates")
            if bbox.get("type") == "Polygon":
                bbox = "{}\n{}".format(coords[0][0], coords[0][-2])
            elif bbox.get("type") == "Point":
                bbox = "Centroïde : {}{}".format(coords[0], coords[1])
            else:
                bbox = "Unknown envelope type (no point nor polygon): " + bbox.get(
                    "type"
                )
        else:
            logging.info("Vector dataset without envelope.")
            pass
        ws["U{}".format(idx)] = bbox

        # geometry
        ws["V{}".format(idx)] = md.get("geometry")

        # resolution
        ws["W{}".format(idx)] = md.get("distance")

        # scale
        ws["X{}".format(idx)] = md.get("scale")

        # features objects
        ws["Y{}".format(idx)] = md.get("features")

        # features attributes
        fields = md.get("feature-attributes", None)
        if fields:
            # count
            ws["Z{}".format(idx)] = len(fields)
            # alphabetic list
            fields_cct = sorted(
                [
                    "{0} ({1})".format(field.get("name"), "description" in field.keys())
                    for field in fields
                ]
            )
            ws["AA{}".format(idx)] = " ;\n".join(fields_cct)
        else:
            logging.info("Vector dataset without any feature attribute")
            pass

        # ---- SPECIFICATIONS # -----------------------------------------------
        specs_in = md.get("specifications", [])
        specs_out = []
        for s_in in specs_in:
            spec = {}
            # translate specification conformity
            if s_in.get("conformant"):
                spec["conformity"] = self.tr("quality", "isConform")
            else:
                spec["conformity"] = self.tr("quality", "isNotConform")
            # ensure other fields
            spec["name"] = s_in.get("specification").get("name")
            spec["link"] = s_in.get("specification").get("link")
            # make data human readable
            spec_date = arrow.get(s_in.get("specification").get("published")[:19])
            spec_date = "{0}".format(spec_date.format(self.dates_fmt, self.locale_fmt))
            spec["date"] = spec_date
            # store into the final list
            specs_out.append(
                "{} {} {} - {}".format(
                    spec.get("name"),
                    spec.get("date"),
                    spec.get("link"),
                    spec.get("conformity"),
                )
            )
        ws["AB{}".format(idx)] = " ;\n".join(specs_out)

        # topology
        ws["AC{}".format(idx)] = md.get("topologicalConsistency", "")

        # ---- CGUs # --------------------------------------------------------
        cgus_in = md.get("conditions", [])
        cgus_out = []
        for c_in in cgus_in:
            cgu = {}
            # ensure other fields
            cgu["description"] = self.clean_xml(c_in.get("description", ""))
            if "license" in c_in.keys():
                cgu["name"] = self.clean_xml(c_in.get("license").get("name", "NR"))
                cgu["link"] = c_in.get("license").get("link", "")
                cgu["content"] = self.clean_xml(c_in.get("license").get("content", ""))
            else:
                cgu["name"] = self.tr("conditions", "noLicense")

            # store into the final list
            cgus_out.append(
                "{} {}. {} {}".format(
                    cgu.get("name"),
                    cgu.get("description", ""),
                    cgu.get("content", ""),
                    cgu.get("link", ""),
                )
            )
        ws["AD{}".format(idx)] = " ;\n".join(cgus_out)

        # ---- LIMITATIONS # -------------------------------------------------
        lims_in = md.get("limitations", [])
        lims_out = []
        for l_in in lims_in:
            limitation = {}
            # ensure other fields
            limitation["description"] = self.clean_xml(l_in.get("description", ""))
            limitation["type"] = self.tr("limitations", l_in.get("type"))
            # legal type
            if l_in.get("type") == "legal":
                limitation["restriction"] = self.tr(
                    "restrictions", l_in.get("restriction")
                )
            else:
                pass
            # INSPIRE precision
            if "directive" in l_in.keys():
                limitation["inspire"] = self.clean_xml(
                    l_in.get("directive").get("name")
                )
                limitation["content"] = self.clean_xml(
                    l_in.get("directive").get("description")
                )
            else:
                pass

            # store into the final list
            lims_out.append(
                "{} {}. {} {} {}".format(
                    limitation.get("type"),
                    limitation.get("description", ""),
                    limitation.get("restriction", ""),
                    limitation.get("content", ""),
                    limitation.get("inspire", ""),
                )
            )

        ws["AE{}".format(idx)] = " ;\n".join(lims_out)

        # CONTACTS
        contacts = md.get("contacts")
        if len(contacts):
            contacts_pt_cct = [
                "{0} ({1})".format(
                    contact.get("contact").get("name"),
                    contact.get("contact").get("email"),
                )
                for contact in contacts
                if contact.get("role") == "pointOfContact"
            ]
            contacts_other_cct = [
                "{0} ({1})".format(
                    contact.get("contact").get("name"),
                    contact.get("contact").get("email"),
                )
                for contact in contacts
                if contact.get("role") != "pointOfContact"
            ]
            ws["AF{}".format(idx)] = len(contacts)
            ws["AG{}".format(idx)] = " ;\n".join(contacts_pt_cct)
            ws["AH{}".format(idx)] = " ;\n".join(contacts_other_cct)
        else:
            ws["AF{}".format(idx)] = 0
            logging.info("Vector dataset without any contact")

        # ACTIONS
        ws["AI{}".format(idx)] = "action:download" in tags
        ws["AJ{}".format(idx)] = "action:view" in tags
        ws["AK{}".format(idx)] = "action:other" in tags

        # LINKS
        link_edit = r'=HYPERLINK("{0}","{1}")'.format(
            "https://app.isogeo.com/resources/" + md.get("_id"), "Editer"
        )
        ws["AL{}".format(idx)] = link_edit
        ws["AL{}".format(idx)].style = "Hyperlink"

        link_visu = r'=HYPERLINK("{0}","{1}")'.format(
            self.url_base + "/m/" + md.get("_id"), "Version en ligne"
        )

        ws["AM{}".format(idx)] = link_visu
        ws["AM{}".format(idx)].style = "Hyperlink"

        # METADATA
        # id
        ws["AN{}".format(idx)] = md.get("_id")

        # creation
        md_created = arrow.get(md.get("_created")[:19])
        md_created = "{0} ({1})".format(
            md_created.format("DD/MM/YYYY", "fr_FR"),
            md_created.humanize(locale="fr_FR"),
        )
        ws["AO{}".format(idx)] = md_created

        # last update
        md_updated = arrow.get(md.get("_modified")[:19])
        md_updated = "{0} ({1})".format(
            md_updated.format("DD/MM/YYYY", "fr_FR"),
            md_updated.humanize(locale="fr_FR"),
        )
        ws["AP{}".format(idx)] = md_updated

        # lang
        ws["AQ{}".format(idx)] = md.get("language")

        # STYLING
        ws["C{}".format(idx)].style = "wrap"
        ws["F{}".format(idx)].style = "wrap"
        ws["G{}".format(idx)].style = "wrap"
        ws["I{}".format(idx)].style = "wrap"
        ws["J{}".format(idx)].style = "wrap"
        ws["K{}".format(idx)].style = "date"
        ws["L{}".format(idx)].style = "date"
        ws["U{}".format(idx)].style = "wrap"
        ws["AA{}".format(idx)].style = "wrap"
        ws["AB{}".format(idx)].style = "wrap"
        ws["AC{}".format(idx)].style = "wrap"
        ws["AD{}".format(idx)].style = "wrap"
        ws["AE{}".format(idx)].style = "wrap"
        ws["AG{}".format(idx)].style = "wrap"
        ws["AH{}".format(idx)].style = "wrap"

        # LOG
        logging.info(
            "Vector metadata stored: {} ({})".format(md.get("name"), md.get("_id"))
        )

        # end of method
        return

    def store_md_raster(self, md, ws, idx):
        """ TO DOCUMENT
        """
        # variables
        tags = md.get("tags")

        ws["A{}".format(idx)] = md.get("title")
        ws["B{}".format(idx)] = md.get("name")
        ws["C{}".format(idx)] = md.get("abstract")

        # path to source
        src_path = md.get("path", "")
        if path.isfile(src_path):
            link_path = r'=HYPERLINK("{0}","{1}")'.format(
                path.dirname(src_path), src_path
            )
            ws["D{}".format(idx)] = link_path
            logging.info("Path reachable")
        else:
            ws["D{}".format(idx)] = src_path
            logging.info("Path not recognized nor reachable")
            pass

        # owner
        ws["E{}".format(idx)] = next(v for k, v in tags.items() if "owner:" in k)
        # KEYWORDS & INSPIRE THEMES
        keywords = []
        inspire = []
        if "keywords" in md.keys():
            for k in md.get("keywords"):
                if k.get("_tag").startswith("keyword:is"):
                    keywords.append(k.get("text"))
                elif k.get("_tag").startswith("keyword:in"):
                    inspire.append(k.get("text"))
                else:
                    logging.info("Unknown keyword type: " + k.get("_tag"))
                    continue
            ws["F{}".format(idx)] = " ;\n".join(sorted(keywords))
            ws["G{}".format(idx)] = " ;\n".join(sorted(inspire))
        else:
            logging.info("Vector dataset without any keyword or INSPIRE theme")

        # conformity
        ws["H{}".format(idx)] = "conformity:inspire" in tags

        # HISTORY
        ws["I{}".format(idx)] = md.get("collectionContext", "")
        ws["J{}".format(idx)] = md.get("collectionMethod", "")

        # validity
        if md.get("validFrom"):
            valid_start = arrow.get(md.get("validFrom"))
            valid_start = "{0}".format(valid_start.format("DD/MM/YYYY", "fr_FR"))
        else:
            valid_start = ""
        ws["K{}".format(idx)] = valid_start

        if md.get("validTo"):
            valid_end = arrow.get(md.get("validTo"))
            valid_end = "{0}".format(valid_end.format("DD/MM/YYYY", "fr_FR"))
        else:
            valid_end = ""
        ws["L{}".format(idx)] = valid_end

        ws["M{}".format(idx)] = md.get("updateFrequency", "")
        ws["N{}".format(idx)] = md.get("validComment", "")

        # EVENTS
        # data creation date
        if md.get("created"):
            data_created = arrow.get(md.get("created"))
            data_created = "{0} ({1})".format(
                data_created.format("DD/MM/YYYY", "fr_FR"),
                data_created.humanize(locale="fr_FR"),
            )
        else:
            data_created = ""
        ws["O{}".format(idx)] = data_created

        # events count
        ws["P{}".format(idx)] = len(md.get("events", ""))

        # data last update
        if md.get("modified"):
            data_updated = arrow.get(md.get("created"))
            data_updated = "{0} ({1})".format(
                data_updated.format("DD/MM/YYYY", "fr_FR"),
                data_updated.humanize(locale="fr_FR"),
            )
        else:
            data_updated = ""
        ws["Q{}".format(idx)] = data_updated

        # TECHNICAL
        # format
        if "format" in md.keys():
            format_lbl = next(v for k, v in tags.items() if "format:" in k)
        else:
            format_lbl = "NR"
        ws["S{}".format(idx)] = "{0} ({1} - {2})".format(
            format_lbl, md.get("formatVersion", "NR"), md.get("encoding", "NR")
        )

        # SRS
        srs = md.setdefault("coordinate-system", {"name": "NR", "code": "NR"})
        ws["T{}".format(idx)] = "{0} ({1})".format(
            srs.get("name", "NR"), srs.get("code", "NR")
        )

        # bounding box
        bbox = md.get("envelope", None)
        if bbox:
            coords = bbox.get("coordinates")
            if bbox.get("type") == "Polygon":
                bbox = "{}\n{}".format(coords[0][0], coords[0][-2])
            elif bbox.get("type") == "Point":
                bbox = "Centroïde : {}{}".format(coords[0], coords[1])
            else:
                bbox = "Unknown envelope type (no point nor polygon): " + bbox.get(
                    "type"
                )
        else:
            logging.info("Vector dataset without envelope.")
            pass
        ws["U{}".format(idx)] = bbox

        # resolution
        ws["V{}".format(idx)] = md.get("distance")

        # scale
        ws["W{}".format(idx)] = md.get("scale")

        # ---- SPECIFICATIONS # -----------------------------------------------
        specs_in = md.get("specifications", [])
        specs_out = []
        for s_in in specs_in:
            spec = {}
            # translate specification conformity
            if s_in.get("conformant"):
                spec["conformity"] = self.tr("quality", "isConform")
            else:
                spec["conformity"] = self.tr("quality", "isNotConform")
            # ensure other fields
            spec["name"] = s_in.get("specification").get("name")
            spec["link"] = s_in.get("specification").get("link")
            # make data human readable
            spec_date = arrow.get(s_in.get("specification").get("published")[:19])
            spec_date = "{0}".format(spec_date.format(self.dates_fmt, self.locale_fmt))
            spec["date"] = spec_date
            # store into the final list
            specs_out.append(
                "{} {} {} - {}".format(
                    spec.get("name"),
                    spec.get("date"),
                    spec.get("link"),
                    spec.get("conformity"),
                )
            )
        ws["X{}".format(idx)] = " ;\n".join(specs_out)
        # topology
        ws["Y{}".format(idx)] = md.get("topologicalConsistency", "")

        # ---- CGUs # --------------------------------------------------------
        cgus_in = md.get("conditions", [])
        cgus_out = []
        for c_in in cgus_in:
            cgu = {}
            # ensure other fields
            cgu["description"] = self.clean_xml(c_in.get("description", ""))
            if "license" in c_in.keys():
                cgu["name"] = self.clean_xml(c_in.get("license").get("name", "NR"))
                cgu["link"] = c_in.get("license").get("link", "")
                cgu["content"] = self.clean_xml(c_in.get("license").get("content", ""))
            else:
                cgu["name"] = self.tr("conditions", "noLicense")

            # store into the final list
            cgus_out.append(
                "{} {}. {} {}".format(
                    cgu.get("name"),
                    cgu.get("description", ""),
                    cgu.get("content", ""),
                    cgu.get("link", ""),
                )
            )
        ws["Z{}".format(idx)] = " ;\n".join(cgus_out)

        # ---- LIMITATIONS # -------------------------------------------------
        lims_in = md.get("limitations", [])
        lims_out = []
        for l_in in lims_in:
            limitation = {}
            # ensure other fields
            limitation["description"] = self.clean_xml(l_in.get("description", ""))
            limitation["type"] = self.tr("limitations", l_in.get("type"))
            # legal type
            if l_in.get("type") == "legal":
                limitation["restriction"] = self.tr(
                    "restrictions", l_in.get("restriction")
                )
            else:
                pass
            # INSPIRE precision
            if "directive" in l_in.keys():
                limitation["inspire"] = self.clean_xml(
                    l_in.get("directive").get("name")
                )
                limitation["content"] = self.clean_xml(
                    l_in.get("directive").get("description")
                )
            else:
                pass

            # store into the final list
            lims_out.append(
                "{} {}. {} {} {}".format(
                    limitation.get("type"),
                    limitation.get("description", ""),
                    limitation.get("restriction", ""),
                    limitation.get("content", ""),
                    limitation.get("inspire", ""),
                )
            )

        ws["AA{}".format(idx)] = " ;\n".join(lims_out)

        # CONTACTS
        contacts = md.get("contacts")
        if len(contacts):
            contacts_pt_cct = [
                "{0} ({1})".format(
                    contact.get("contact").get("name"),
                    contact.get("contact").get("email"),
                )
                for contact in contacts
                if contact.get("role") == "pointOfContact"
            ]
            contacts_other_cct = [
                "{0} ({1})".format(
                    contact.get("contact").get("name"),
                    contact.get("contact").get("email"),
                )
                for contact in contacts
                if contact.get("role") != "pointOfContact"
            ]
            ws["AB{}".format(idx)] = len(contacts)
            ws["AC{}".format(idx)] = " ;\n".join(contacts_pt_cct)
            ws["AD{}".format(idx)] = " ;\n".join(contacts_other_cct)
        else:
            ws["AB{}".format(idx)] = 0
            logging.info("Vector dataset without any contact")

        # ACTIONS
        ws["AE{}".format(idx)] = "action:download" in tags
        ws["AF{}".format(idx)] = "action:view" in tags
        ws["AG{}".format(idx)] = "action:other" in tags

        # LINKS
        link_edit = r'=HYPERLINK("{0}","{1}")'.format(
            "https://app.isogeo.com/resources/" + md.get("_id"), "Editer"
        )
        ws["AH{}".format(idx)] = link_edit
        ws["AH{}".format(idx)].style = "Hyperlink"

        link_visu = r'=HYPERLINK("{0}","{1}")'.format(
            self.url_base + "/m/" + md.get("_id"), "Version en ligne"
        )

        ws["AI{}".format(idx)] = link_visu
        ws["AI{}".format(idx)].style = "Hyperlink"
        # METADATA
        # id
        ws["AJ{}".format(idx)] = md.get("_id")

        # creation
        md_created = arrow.get(md.get("_created")[:19])
        md_created = "{0} ({1})".format(
            md_created.format("DD/MM/YYYY", "fr_FR"),
            md_created.humanize(locale="fr_FR"),
        )
        ws["AK{}".format(idx)] = md_created

        # last update
        md_updated = arrow.get(md.get("_modified")[:19])
        md_updated = "{0} ({1})".format(
            md_updated.format("DD/MM/YYYY", "fr_FR"),
            md_updated.humanize(locale="fr_FR"),
        )
        ws["AL{}".format(idx)] = md_updated

        # lang
        ws["AM{}".format(idx)] = md.get("language")

        # STYLING
        ws["C{}".format(idx)].style = "wrap"
        ws["F{}".format(idx)].style = "wrap"
        ws["G{}".format(idx)].style = "wrap"
        ws["I{}".format(idx)].style = "wrap"
        ws["J{}".format(idx)].style = "wrap"
        ws["K{}".format(idx)].style = "date"
        ws["L{}".format(idx)].style = "date"
        ws["U{}".format(idx)].style = "wrap"
        ws["X{}".format(idx)].style = "wrap"
        ws["Y{}".format(idx)].style = "wrap"
        ws["Z{}".format(idx)].style = "wrap"
        ws["AA{}".format(idx)].style = "wrap"
        ws["AC{}".format(idx)].style = "wrap"
        ws["AD{}".format(idx)].style = "wrap"

        # LOG
        logging.info(
            "Raster metadata stored: {} ({})".format(md.get("name"), md.get("_id"))
        )

        # end of method
        return

    def store_md_service(self, md, ws, idx):
        """ TO DOCUMENT
        """
        # variables
        tags = md.get("tags")

        ws["A{}".format(idx)] = md.get("title")
        ws["B{}".format(idx)] = md.get("name")
        ws["C{}".format(idx)] = md.get("abstract")

        # path of GetCapabilities
        src_path = md.get("path", None)
        if src_path:
            link_path = r'=HYPERLINK("{0}","{1}")'.format(src_path, src_path)
            ws["D{}".format(idx)] = link_path
        else:
            logging.info("GetCapabilities missing")
            pass

        # owner
        ws["E{}".format(idx)] = next(v for k, v in tags.items() if "owner:" in k)
        # KEYWORDS
        if "keywords" in md.keys():
            keywords = [
                k.get("text")
                for k in md.get("keywords")
                if k.get("_tag").startswith("keyword:is")
            ]
            ws["F{}".format(idx)] = " ;\n".join(sorted(keywords))
        else:
            logging.info("Service without any keyword")

        # conformity
        ws["G{}".format(idx)] = "conformity:inspire" in tags

        # EVENTS
        # data creation date
        if md.get("created"):
            data_created = arrow.get(md.get("created"))
            data_created = "{0} ({1})".format(
                data_created.format("DD/MM/YYYY", "fr_FR"),
                data_created.humanize(locale="fr_FR"),
            )
        else:
            data_created = ""
        ws["H{}".format(idx)] = data_created

        # events count
        ws["I{}".format(idx)] = len(md.get("events", ""))

        # data last update
        if md.get("modified"):
            data_updated = arrow.get(md.get("created"))
            data_updated = "{0} ({1})".format(
                data_updated.format("DD/MM/YYYY", "fr_FR"),
                data_updated.humanize(locale="fr_FR"),
            )
        else:
            data_updated = ""
        ws["J{}".format(idx)] = data_updated

        # TECHNICAL
        # format
        if "format" in md.keys():
            format_lbl = next(v for k, v in tags.items() if "format:" in k)
        else:
            format_lbl = "NR"
        ws["L{}".format(idx)] = "{0} ({1})".format(
            format_lbl, md.get("formatVersion", "NR")
        )

        # bounding box
        bbox = md.get("envelope", None)
        if bbox:
            coords = bbox.get("coordinates")
            if bbox.get("type") == "Polygon":
                bbox = "{}\n{}".format(coords[0][0], coords[0][-2])
            elif bbox.get("type") == "Point":
                bbox = "Point unique : {}{}".format(coords[0], coords[1])
            elif bbox.get("type") == "Point":
                bbox = "Ligne unique : {}{}".format(coords[0], coords[1])
            else:
                bbox = "Unknown envelope type (no point nor polygon): " + bbox.get(
                    "type"
                )
        else:
            logging.info("Vector dataset without envelope.")
            pass
        ws["M{}".format(idx)] = bbox

        # ---- SPECIFICATIONS # -----------------------------------------------
        specs_in = md.get("specifications", [])
        specs_out = []
        for s_in in specs_in:
            spec = {}
            # translate specification conformity
            if s_in.get("conformant"):
                spec["conformity"] = self.tr("quality", "isConform")
            else:
                spec["conformity"] = self.tr("quality", "isNotConform")
            # ensure other fields
            spec["name"] = s_in.get("specification").get("name")
            spec["link"] = s_in.get("specification").get("link")
            # make data human readable
            spec_date = arrow.get(s_in.get("specification").get("published")[:19])
            spec_date = "{0}".format(spec_date.format(self.dates_fmt, self.locale_fmt))
            spec["date"] = spec_date
            # store into the final list
            specs_out.append(
                "{} {} {} - {}".format(
                    spec.get("name"),
                    spec.get("date"),
                    spec.get("link"),
                    spec.get("conformity"),
                )
            )
        ws["N{}".format(idx)] = " ;\n".join(specs_out)

        # ---- CGUs # --------------------------------------------------------
        cgus_in = md.get("conditions", [])
        cgus_out = []
        for c_in in cgus_in:
            cgu = {}
            # ensure other fields
            cgu["description"] = self.clean_xml(c_in.get("description", ""))
            if "license" in c_in.keys():
                cgu["name"] = self.clean_xml(c_in.get("license").get("name", "NR"))
                cgu["link"] = c_in.get("license").get("link", "")
                cgu["content"] = self.clean_xml(c_in.get("license").get("content", ""))
            else:
                cgu["name"] = self.tr("conditions", "noLicense")

            # store into the final list
            cgus_out.append(
                "{} {}. {} {}".format(
                    cgu.get("name"),
                    cgu.get("description", ""),
                    cgu.get("content", ""),
                    cgu.get("link", ""),
                )
            )
        ws["O{}".format(idx)] = " ;\n".join(cgus_out)

        # ---- LIMITATIONS # -------------------------------------------------
        lims_in = md.get("limitations", [])
        lims_out = []
        for l_in in lims_in:
            limitation = {}
            # ensure other fields
            limitation["description"] = self.clean_xml(l_in.get("description", ""))
            limitation["type"] = self.tr("limitations", l_in.get("type"))
            # legal type
            if l_in.get("type") == "legal":
                limitation["restriction"] = self.tr(
                    "restrictions", l_in.get("restriction")
                )
            else:
                pass
            # INSPIRE precision
            if "directive" in l_in.keys():
                limitation["inspire"] = self.clean_xml(
                    l_in.get("directive").get("name")
                )
                limitation["content"] = self.clean_xml(
                    l_in.get("directive").get("description")
                )
            else:
                pass

            # store into the final list
            lims_out.append(
                "{} {}. {} {} {}".format(
                    limitation.get("type"),
                    limitation.get("description", ""),
                    limitation.get("restriction", ""),
                    limitation.get("content", ""),
                    limitation.get("inspire", ""),
                )
            )

        ws["P{}".format(idx)] = " ;\n".join(lims_out)

        # CONTACTS
        contacts = md.get("contacts")
        if len(contacts):
            contacts_pt_cct = [
                "{0} ({1})".format(
                    contact.get("contact").get("name"),
                    contact.get("contact").get("email"),
                )
                for contact in contacts
                if contact.get("role") == "pointOfContact"
            ]
            contacts_other_cct = [
                "{0} ({1})".format(
                    contact.get("contact").get("name"),
                    contact.get("contact").get("email"),
                )
                for contact in contacts
                if contact.get("role") != "pointOfContact"
            ]
            ws["Q{}".format(idx)] = len(contacts)
            ws["R{}".format(idx)] = " ;\n".join(contacts_pt_cct)
            ws["S{}".format(idx)] = " ;\n".join(contacts_other_cct)
        else:
            ws["Q{}".format(idx)] = 0
            logging.info("Service without any contact")

        # ACTIONS
        ws["T{}".format(idx)] = "action:download" in tags
        ws["U{}".format(idx)] = "action:view" in tags
        ws["V{}".format(idx)] = "action:other" in tags

        # LINKS
        link_edit = r'=HYPERLINK("{0}","{1}")'.format(
            "https://app.isogeo.com/resources/" + md.get("_id"), "Editer"
        )
        ws["W{}".format(idx)] = link_edit
        ws["W{}".format(idx)].style = "Hyperlink"

        link_visu = r'=HYPERLINK("{0}","{1}")'.format(
            self.url_base + "/m/" + md.get("_id"), "Version en ligne"
        )

        ws["X{}".format(idx)] = link_visu
        ws["X{}".format(idx)].style = "Hyperlink"

        # METADATA
        # id
        ws["Y{}".format(idx)] = md.get("_id")

        # creation
        md_created = arrow.get(md.get("_created")[:19])
        md_created = "{0} ({1})".format(
            md_created.format("DD/MM/YYYY", "fr_FR"),
            md_created.humanize(locale="fr_FR"),
        )
        ws["Z{}".format(idx)] = md_created

        # last update
        md_updated = arrow.get(md.get("_modified")[:19])
        md_updated = "{0} ({1})".format(
            md_updated.format("DD/MM/YYYY", "fr_FR"),
            md_updated.humanize(locale="fr_FR"),
        )
        ws["AA{}".format(idx)] = md_updated

        # lang
        ws["AB{}".format(idx)] = md.get("language")

        # STYLING
        ws["C{}".format(idx)].style = "wrap"
        ws["F{}".format(idx)].style = "wrap"
        ws["M{}".format(idx)].style = "wrap"
        ws["N{}".format(idx)].style = "wrap"
        ws["O{}".format(idx)].style = "wrap"
        ws["P{}".format(idx)].style = "wrap"
        ws["R{}".format(idx)].style = "wrap"
        ws["S{}".format(idx)].style = "wrap"

        # LOG
        logging.info(
            "Service metadata stored: {} ({})".format(md.get("name"), md.get("_id"))
        )

        # end of method
        return

    def store_md_resource(self, md, ws, idx):
        """ TO DOCUMENT
        """
        # variables
        tags = md.get("tags")

        ws["A{}".format(idx)] = md.get("title", "NR")
        ws["B{}".format(idx)] = md.get("abstract", "")
        ws["C{}".format(idx)] = md.get("path", "")
        ws["D{}".format(idx)] = md.get("owner")

        # KEYWORDS
        if "keywords" in md.keys():
            keywords = [
                k.get("text")
                for k in md.get("keywords")
                if k.get("_tag").startswith("keyword:is")
            ]
            ws["E{}".format(idx)] = " ;\n".join(sorted(keywords))
        else:
            logging.info("Service without any keyword")

        # EVENTS
        # data creation date
        if md.get("created"):
            data_created = arrow.get(md.get("created"))
            data_created = "{0} ({1})".format(
                data_created.format("DD/MM/YYYY", "fr_FR"),
                data_created.humanize(locale="fr_FR"),
            )
        else:
            data_created = ""
        ws["F{}".format(idx)] = data_created

        # events count
        ws["G{}".format(idx)] = len(md.get("events", ""))

        # data last update
        if md.get("modified"):
            data_updated = arrow.get(md.get("created"))
            data_updated = "{0} ({1})".format(
                data_updated.format("DD/MM/YYYY", "fr_FR"),
                data_updated.humanize(locale="fr_FR"),
            )
        else:
            data_updated = ""
        ws["H{}".format(idx)] = data_updated

        # TECHNICAL
        # format
        if "format:" in tags.keys():
            format_lbl = next(v for k, v in tags.items() if "format:" in k)
        else:
            format_lbl = "NR"
        ws["J{}".format(idx)] = "{0} ({1} - {2})".format(
            format_lbl, md.get("formatVersion", "NR"), md.get("encoding", "NR")
        )

        # ---- CGUs # --------------------------------------------------------
        cgus_in = md.get("conditions", [])
        cgus_out = []
        for c_in in cgus_in:
            cgu = {}
            # ensure other fields
            cgu["description"] = self.clean_xml(c_in.get("description", ""))
            if "license" in c_in.keys():
                cgu["name"] = self.clean_xml(c_in.get("license").get("name", "NR"))
                cgu["link"] = c_in.get("license").get("link", "")
                cgu["content"] = self.clean_xml(c_in.get("license").get("content", ""))
            else:
                cgu["name"] = self.tr("conditions", "noLicense")

            # store into the final list
            cgus_out.append(
                "{} {}. {} {}".format(
                    cgu.get("name"),
                    cgu.get("description", ""),
                    cgu.get("content", ""),
                    cgu.get("link", ""),
                )
            )
        ws["K{}".format(idx)] = " ;\n".join(cgus_out)

        # ---- LIMITATIONS # -------------------------------------------------
        lims_in = md.get("limitations", [])
        lims_out = []
        for l_in in lims_in:
            limitation = {}
            # ensure other fields
            limitation["description"] = self.clean_xml(l_in.get("description", ""))
            limitation["type"] = self.tr("limitations", l_in.get("type"))
            # legal type
            if l_in.get("type") == "legal":
                limitation["restriction"] = self.tr(
                    "restrictions", l_in.get("restriction")
                )
            else:
                pass
            # INSPIRE precision
            if "directive" in l_in.keys():
                limitation["inspire"] = self.clean_xml(
                    l_in.get("directive").get("name")
                )
                limitation["content"] = self.clean_xml(
                    l_in.get("directive").get("description")
                )
            else:
                pass

            # store into the final list
            lims_out.append(
                "{} {}. {} {} {}".format(
                    limitation.get("type"),
                    limitation.get("description", ""),
                    limitation.get("restriction", ""),
                    limitation.get("content", ""),
                    limitation.get("inspire", ""),
                )
            )

        ws["L{}".format(idx)] = " ;\n".join(lims_out)

        # CONTACTS
        contacts = md.get("contacts")
        if len(contacts):
            contacts_pt_cct = [
                "{0} ({1})".format(
                    contact.get("contact").get("name"),
                    contact.get("contact").get("email"),
                )
                for contact in contacts
                if contact.get("role") == "pointOfContact"
            ]
            contacts_other_cct = [
                "{0} ({1})".format(
                    contact.get("contact").get("name"),
                    contact.get("contact").get("email"),
                )
                for contact in contacts
                if contact.get("role") != "pointOfContact"
            ]
            ws["M{}".format(idx)] = len(contacts)
            ws["N{}".format(idx)] = " ;\n".join(contacts_pt_cct)
            ws["O{}".format(idx)] = " ;\n".join(contacts_other_cct)
        else:
            ws["M{}".format(idx)] = 0
            logging.info("Service without any contact")

        # ACTIONS
        ws["P{}".format(idx)] = "action:download" in tags
        ws["Q{}".format(idx)] = "action:view" in tags
        ws["R{}".format(idx)] = "action:other" in tags

        # LINKS
        link_edit = r'=HYPERLINK("{0}","{1}")'.format(
            "https://app.isogeo.com/resources/" + md.get("_id"), "Editer"
        )
        ws["S{}".format(idx)] = link_edit
        ws["S{}".format(idx)].style = "Hyperlink"

        link_visu = r'=HYPERLINK("{0}","{1}")'.format(
            self.url_base + "/m/" + md.get("_id"), "Version en ligne"
        )

        ws["T{}".format(idx)] = link_visu
        ws["T{}".format(idx)].style = "Hyperlink"

        # METADATA
        # id
        ws["U{}".format(idx)] = md.get("_id")

        # creation
        md_created = arrow.get(md.get("_created")[:19])
        md_created = "{0} ({1})".format(
            md_created.format("DD/MM/YYYY", "fr_FR"),
            md_created.humanize(locale="fr_FR"),
        )
        ws["V{}".format(idx)] = md_created

        # last update
        md_updated = arrow.get(md.get("_modified")[:19])
        md_updated = "{0} ({1})".format(
            md_updated.format("DD/MM/YYYY", "fr_FR"),
            md_updated.humanize(locale="fr_FR"),
        )
        ws["W{}".format(idx)] = md_updated

        # lang
        ws["X{}".format(idx)] = md.get("language")

        # STYLING
        ws["C{}".format(idx)].style = "wrap"
        ws["F{}".format(idx)].style = "wrap"
        ws["M{}".format(idx)].style = "wrap"
        ws["N{}".format(idx)].style = "wrap"
        ws["O{}".format(idx)].style = "wrap"
        ws["P{}".format(idx)].style = "wrap"
        ws["R{}".format(idx)].style = "wrap"
        ws["S{}".format(idx)].style = "wrap"

        # LOG
        logging.info(
            "Resource metadata stored: {} ({})".format(md.get("name"), md.get("_id"))
        )

        # end of method
        return

    # ------------ Writing metadata ---------------------

    def tunning_worksheets(self):
        """ CLEAN UP & TUNNING
        """
        for sheet in self.worksheets:
            # Freezing panes
            c_freezed = sheet["B2"]
            sheet.freeze_panes = c_freezed

            # Print properties
            sheet.print_options.horizontalCentered = True
            sheet.print_options.verticalCentered = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

            # Others properties
            wsprops = sheet.sheet_properties
            wsprops.filterMode = True

            # enable filters
            sheet.auto_filter.ref = str("A1:{}{}").format(
                get_column_letter(sheet.max_column), sheet.max_row
            )
        pass

    def remove_accents(self, input_str, substitute=""):
        """Clean string from special characters.

        source: http://stackoverflow.com/a/5843560
        """
        return unicode(substitute).join(char for char in input_str if char.isalnum())

    def clean_xml(self, invalid_xml, mode="soft", substitute="_"):
        """Clean string of XML invalid characters.

        source: http://stackoverflow.com/a/13322581/2556577
        """
        # assumptions:
        #   doc = *( start_tag / end_tag / text )
        #   start_tag = '<' name *attr [ '/' ] '>'
        #   end_tag = '<' '/' name '>'
        ws = r"[ \t\r\n]*"  # allow ws between any token
        name = "[a-zA-Z]+"  # note: expand if necessary but the stricter the better
        attr = '{name} {ws} = {ws} "[^"]*"'  # note: fragile against missing '"'; no "'"
        start_tag = "< {ws} {name} {ws} (?:{attr} {ws})* /? {ws} >"
        end_tag = "{ws}".join(["<", "/", "{name}", ">"])
        tag = "{start_tag} | {end_tag}"

        assert "{{" not in tag
        while "{" in tag:  # unwrap definitions
            tag = tag.format(**vars())

        tag_regex = re.compile("(%s)" % tag, flags=re.VERBOSE)

        # escape &, <, > in the text
        iters = [iter(tag_regex.split(invalid_xml))] * 2
        pairs = izip_longest(*iters, fillvalue="")  # iterate 2 items at a time

        # get the clean version
        clean_version = "".join(escape(text) + tag for text, tag in pairs)
        if mode == "strict":
            clean_version = re.sub(r"<.*?>", substitute, clean_version)
        else:
            pass
        return clean_version


# #############################################################################
# ##### Stand alone program ########
# ##################################

if __name__ == "__main__":
    """ Standalone execution and tests
    """
    # ------------ Specific imports ---------------------
    from ConfigParser import SafeConfigParser  # to manage options.ini

    # ------------ Settings from ini file ----------------
    if not path.isfile(path.realpath(r"..\settings.ini")):
        logging.error(
            "To execute this script as standalone, you need to store your Isogeo application settings in a isogeo_params.ini file. You can use the template to set your own."
        )
        raise ValueError("isogeo_params.ini file missing.")
    else:
        pass

    config = SafeConfigParser()
    config.read(r"..\settings_dev.ini")

    settings = {s: dict(config.items(s)) for s in config.sections()}
    app_id = settings.get("auth").get("app_id")
    app_secret = settings.get("auth").get("app_secret")
    client_lang = settings.get("basics").get("def_codelang")
    def_oc = settings.get("basics").get(
        "def_oc",
        "http://open.isogeo.com/s/fbc04d8809784283b60e5f78ddc1e2cb/LoXCxjcjvRclWdCPzh17HXfhNYO20",
    )

    # ------------ Connecting to Isogeo API ----------------
    # instanciating the class
    isogeo = Isogeo(client_id=app_id, client_secret=app_secret, lang="fr")

    token = isogeo.connect()

    # ------------ Isogeo search --------------------------
    includes = [
        "conditions",
        "contacts",
        "coordinate-system",
        "events",
        "feature-attributes",
        "keywords",
        "limitations",
        "links",
        "specifications",
    ]

    search_results = isogeo.search(token, sub_resources=includes)

    # ------------ REAL START ----------------------------
    wb = Isogeo2xlsx(url_base=def_oc)
    wb.set_worksheets(auto=search_results.get("tags").keys())

    # parsing metadata
    for md in search_results.get("results"):
        wb.store_metadatas(md)

    # tunning
    wb.tunning_worksheets()

    # saving the test file
    dstamp = datetime.now()
    wb.save(
        r"..\output\test_isogeo2xlsx_{0}{1}{2}{3}{4}{5}.xlsx".format(
            dstamp.year,
            dstamp.month,
            dstamp.day,
            dstamp.hour,
            dstamp.minute,
            dstamp.second,
        )
    )


### DEV NOTES
# http://wiki.openstreetmap.org/wiki/FR:Parcourir#URL_avec_bbox
# http://wiki.openstreetmap.org/wiki/Layer_URL_parameter
# https://www.openstreetmap.org/?bbox=22.3418234%2C57.5129102%2C22.5739625%2C57.6287332&layers=H&box=yes
