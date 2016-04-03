# -*- coding: UTF-8 -*-
#!/usr/bin/env python
from __future__ import (absolute_import, print_function, unicode_literals)
# ------------------------------------------------------------------------------
# Name:         Isogeo to Microsoft Excel 2010
# Purpose:      Get metadatas from an Isogeo share and store it into
#               a Excel worksheet. It's one of the submodules of
#               isogeo2office (https://bitbucket.org/isogeo/isogeo-2-office).
#
# Author:       Julien Moura (@geojulien)
#
# Python:       2.7.x
# Created:      14/08/2014
# Updated:      30/01/2016
# ------------------------------------------------------------------------------

# ##############################################################################
# ########## Libraries #############
# ##################################

# Standard library
import logging
from os import path

# Python 3 backported
from collections import OrderedDict     # ordered dictionary

# 3rd party library
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties
from openpyxl.styles import Style, Font, Alignment

# ##############################################################################
# ########## Classes ###############
# ##################################


class files2xlsx(Workbook):
    """ files2xlsx
    """
    li_cols_vector = [
                      "nomfic",
                      "path",
                      "theme",
                      "num_attrib",
                      "num_objets",
                      "geometrie",
                      "srs",
                      "srs_type",
                      "codepsg",
                      "emprise",
                      "date_crea",
                      "date_actu",
                      "format",
                      "li_depends",
                      "tot_size",
                      "li_chps",
                      "gdal_warn"
                     ]

    li_cols_raster = [
                     "nomfic",
                     "path",
                     "theme",
                     "size_Y",
                     "size_X",
                     "pixel_w",
                     "pixel_h",
                     "origin_x",
                     "origin_y",
                     "srs_type",
                     "codepsg",
                     "emprise",
                     "date_crea",
                     "date_actu",
                     "num_bands",
                     "format",
                     "compression",
                     "coloref",
                     "li_depends",
                     "tot_size",
                     "gdal_warn",
                     ]

    li_cols_filedb = [
                      "nomfic",
                      "path",
                      "theme",
                      "tot_size",
                      "date_crea",
                      "date_actu",
                      "feats_class",
                      "num_attrib",
                      "num_objets",
                      "geometrie",
                      "srs",
                      "srs_type",
                      "codepsg",
                      "emprise",
                      "li_chps",
                      ]

    li_cols_mapdocs = [
                      "nomfic",
                      "path",
                      "theme",
                      "custom_title",
                      "creator_prod",
                      "keywords",
                      "subject",
                      "res_image",
                      "tot_size",
                      "date_crea",
                      "date_actu",
                      "origin_x",
                      "origin_y",
                      "srs",
                      "srs_type",
                      "codepsg",
                      "sub_layers",
                      "num_attrib",
                      "num_objets",
                      "li_chps",
                      ]

    li_cols_caodao = [
                      "nomfic",
                      "path",
                      "theme",
                      "tot_size",
                      "date_crea",
                      "date_actu",
                      "sub_layers",
                      "num_attrib",
                      "num_objets",
                      "geometrie",
                      "srs",
                      "srs_type",
                      "codepsg",
                      "emprise",
                      "li_chps",
                      ]

    li_cols_sgbd = [
                      "nomfic",
                      "conn_chain",
                      "schema",
                      "num_attrib",
                      "num_objets",
                      "geometrie",
                      "srs",
                      "srs_type",
                      "codepsg",
                      "emprise",
                      "date_crea",
                      "date_actu",
                      "format",
                      "li_chps",
                      ]

    def __init__(self, lang="EN", texts=OrderedDict()):
        """ TO DOC

        Keyword arguments:

        """
        super(files2xlsx, self).__init__()
        # super(files2xlsx, self).__init__(write_only=True)
        self.texts = texts

        # styles
        self.s_error = Style(font=Font(color="FF0000"))
        self.s_header = Style(alignment=Alignment(horizontal='center',
                                                  vertical='center'
                                                  ),
                              font=Font(size=12,
                                        bold=True,
                                        )
                              )
        self.s_link = Style(font=Font(underline="single"))
        self.s_wrap = Style(alignment=Alignment(wrap_text=True))

        # deleting the default worksheet
        ws = self.active
        self.remove_sheet(ws)

    # ------------ Setting workbook ---------------------

    def set_worksheets(self, has_vector=0, has_raster=0, has_filedb=0,
                       has_mapdocs=0, has_cad=0, has_sgbd=0):
        """ adds news sheets depending on present metadata types
        """
        # SHEETS & HEADERS
        if has_vector:
            self.ws_v = self.create_sheet(title=self.texts.get("sheet_vectors"))
            # headers
            self.ws_v.append([self.texts.get(i) for i in self.li_cols_vector])
            # styling
            for i in self.li_cols_vector:
                self.ws_v.cell(row=1,
                               column=self.li_cols_vector.index(i) + 1).style = self.s_header

            # initialize line counter
            self.idx_v = 1
        else:
            pass

        if has_raster:
            self.ws_r = self.create_sheet(title=self.texts.get("sheet_rasters"))
            # headers
            self.ws_r.append([self.texts.get(i) for i in self.li_cols_raster])
            # styling
            for i in self.li_cols_raster:
                self.ws_r.cell(row=1,
                               column=self.li_cols_raster.index(i) + 1).style = self.s_header

            # initialize line counter
            self.idx_r = 1
        else:
            pass

        if has_filedb:
            self.ws_fdb = self.create_sheet(title=self.texts.get("sheet_filedb"))
            # headers
            self.ws_fdb.append([self.texts.get(i) for i in self.li_cols_filedb])
            # styling
            for i in self.li_cols_filedb:
                self.ws_fdb.cell(row=1,
                                 column=self.li_cols_filedb.index(i) + 1).style = self.s_header

            # initialize line counter
            self.idx_f = 1
        else:
            pass

        if has_mapdocs:
            self.ws_mdocs = self.create_sheet(title=self.texts.get("sheet_maplans"))
            # headers
            self.ws_mdocs.append([self.texts.get(i) for i in self.li_cols_mapdocs])
            # styling
            for i in self.li_cols_mapdocs:
                self.ws_mdocs.cell(row=1,
                                   column=self.li_cols_mapdocs.index(i) + 1).style = self.s_header

            # initialize line counter
            self.idx_m = 1
        else:
            pass

        if has_cad:
            self.ws_cad = self.create_sheet(title=self.texts.get("sheet_cdao"))
            # headers
            self.ws_cad.append([self.texts.get(i) for i in self.li_cols_caodao])
            # styling
            for i in self.li_cols_caodao:
                self.ws_cad.cell(row=1,
                                 column=self.li_cols_caodao.index(i) + 1).style = self.s_header

            # initialize line counter
            self.idx_c = 1
        else:
            pass

        if has_sgbd:
            self.ws_sgbd = self.create_sheet(title="PostGIS")
            # headers
            self.ws_sgbd.append([self.texts.get(i) for i in self.li_cols_sgbd])
            # styling
            for i in self.li_cols_sgbd:
                self.ws_sgbd.cell(row=1,
                                  column=self.li_cols_sgbd.index(i) + 1).style = self.s_header
            # initialize line counter
            self.idx_s = 1
        else:
            pass

        # end of method
        return

    def tunning_worksheets(self):
        """ CLEAN UP & TUNNING
        """
        for sheet in self.worksheets:
            # Freezing panes
            c_freezed = sheet['B2']
            sheet.freeze_panes = c_freezed

            # Print properties
            sheet.print_options.horizontalCentered = True
            sheet.print_options.verticalCentered = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

            # Others properties
            wsprops = sheet.sheet_properties
            wsprops.filterMode = True

            # enable filters
            sheet.auto_filter.ref = str("A1:{}{}").format(get_column_letter(sheet.max_column),
                                                          sheet.max_row)
        pass

    # ------------ Writing metadata ---------------------
    def store_md_vector(self, layer, fields):
        """ Storing metadata about a vector dataset
        """
        # increment line
        self.idx_v += 1

        # local variables
        champs = ""

        # in case of a source error
        if layer.get('error'):
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.texts.get(layer.get('error'))
            logging.warning('\tproblem detected')
            self.ws_v["A{}".format(self.idx_v)] = layer.get('name')
            self.ws_v["A{}".format(self.idx_v)].style = self.s_error
            link = r'=HYPERLINK("{0}","{1}")'.format(layer.get(u'folder'),
                                                     self.texts.get('browse'))
            self.ws_v["B{}".format(self.idx_v)] = link
            self.ws_v["B{}".format(self.idx_v)].style = self.s_error
            self.ws_v["C{}".format(self.idx_v)] = err_mess
            self.ws_v["C{}".format(self.idx_v)].style = self.s_error
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_v["A{}".format(self.idx_v)] = layer.get('name')

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{0}","{1}")'.format(layer.get(u'folder'),
                                                 self.texts.get('browse'))
        self.ws_v["B{}".format(self.idx_v)] = link
        self.ws_v["B{}".format(self.idx_v)].style = self.s_link

        # Name of parent folder with an exception if this is the format name
        self.ws_v["C{}".format(self.idx_v)] = path.basename(layer.get(u'folder'))

        # Fields count
        self.ws_v["D{}".format(self.idx_v)] = layer.get(u'num_fields')
        # Objects count
        self.ws_v["E{}".format(self.idx_v)] = layer.get(u'num_obj')
        # Geometry type
        self.ws_v["F{}".format(self.idx_v)] = layer.get(u'type_geom')
        # Name of srs
        self.ws_v["G{}".format(self.idx_v)] = layer.get(u'srs')
        # Type of SRS
        self.ws_v["H{}".format(self.idx_v)] = layer.get(u'srs_type')
        # EPSG code
        self.ws_v["I{}".format(self.idx_v)] = layer.get(u'EPSG')
        # Spatial extent
        emprise = "Xmin : {0} - Xmax : {1} | \nYmin : {2} - Ymax : {3}"\
                  .format(unicode(layer.get(u'Xmin')),
                          unicode(layer.get(u'Xmax')),
                          unicode(layer.get(u'Ymin')),
                          unicode(layer.get(u'Ymax')))
        self.ws_v["J{}".format(self.idx_v)].style = self.s_wrap
        self.ws_v["J{}".format(self.idx_v)] = emprise

        # Creation date
        self.ws_v["K{}".format(self.idx_v)] = layer.get(u'date_crea')
        # Last update date
        self.ws_v["L{}".format(self.idx_v)] = layer.get(u'date_actu')
        # Format of data
        self.ws_v["M{}".format(self.idx_v)] = layer.get(u'type')
        # dependencies
        self.ws_v["N{}".format(self.idx_v)].style.alignment.wrap_text = True
        self.ws_v.cell("N{}".format(self.idx_v)).value = u' |\n '.join(layer.get(u'dependencies'))
        # total size
        self.ws_v["O{}".format(self.idx_v)] = layer.get(u'total_size')

        # Field informations
        for chp in fields.keys():
            # field type
            if 'Integer' in fields[chp][0]:
                tipo = self.texts.get(u'entier')
            elif fields[chp][0] == 'Real':
                tipo = self.texts.get(u'reel')
            elif fields[chp][0] == 'String':
                tipo = self.texts.get(u'string')
            elif fields[chp][0] == 'Date':
                tipo = self.texts.get(u'date')
            else:
                tipo = "unknown"
                logging.warning(chp, " unknown type")

            # concatenation of field informations
            try:
                champs = champs + chp +\
                          u" (" + tipo + self.texts.get(u'longueur') +\
                          unicode(fields[chp][1]) +\
                          self.texts.get(u'precision') +\
                          unicode(fields[chp][2]) + u") ; "
            except UnicodeDecodeError:
                logging.warning('Field name with special letters: {}'.format(chp.decode('latin1')))
                # decode the fucking field name
                champs = champs + chp.decode('latin1') \
                + u" ({}, Lg. = {}, Pr. = {}) ;".format(tipo,
                                                        fields[chp][1],
                                                        fields[chp][2])
                # then continue
                continue

        # Once all fieds explored, write them
        self.ws_v["P{}".format(self.idx_v)] = champs

        # in case of a source error
        if layer.get('err_gdal')[0] != 0:
            logging.warning('\tproblem detected')
            self.ws_v["Q{}".format(self.idx_v)] = "{0} : {1}".format(layer.get('err_gdal')[0],
                                                                     layer.get('err_gdal')[1])
            self.ws_v["Q{}".format(self.idx_v)].style = self.s_error
        else:
            pass

        # end of method
        return

    def store_md_raster(self, layer, bands):
        """ Storing metadata about a raster dataset
        """
        # increment line
        self.idx_r += 1

        # in case of a source error
        if layer.get('error'):
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.texts.get(layer.get('error'))
            logging.warning('\tproblem detected')
            self.ws_r["A{}".format(self.idx_r)] = layer.get('name')
            link = r'=HYPERLINK("{0}","{1}")'.format(layer.get(u'folder'),
                                                     self.texts.get('browse'))
            self.ws_r["B{}".format(self.idx_r)] = link
            self.ws_r["B{}".format(self.idx_r)].style = self.s_error
            self.ws_r["C{}".format(self.idx_r)] = err_mess
            self.ws_r["C{}".format(self.idx_r)].style = self.s_error
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_r["A{}".format(self.idx_r)] = layer.get('name')

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{0}","{1}")'.format(layer.get(u'folder'),
                                                 self.texts.get('browse'))
        self.ws_r["B{}".format(self.idx_r)] = link
        self.ws_r["B{}".format(self.idx_r)].style = self.s_link

        # Name of parent folder with an exception if this is the format name
        self.ws_r["C{}".format(self.idx_r)] = path.basename(layer.get(u'folder'))

        # Image dimensions
        self.ws_r["D{}".format(self.idx_r)] = layer.get(u'num_rows')
        self.ws_r["E{}".format(self.idx_r)] = layer.get(u'num_cols')

        # Pixel dimensions
        self.ws_r["F{}".format(self.idx_r)] = layer.get(u'pixelWidth')
        self.ws_r["G{}".format(self.idx_r)] = layer.get(u'pixelHeight')

        # Image dimensions
        self.ws_r["H{}".format(self.idx_r)] = layer.get(u'xOrigin')
        self.ws_r["I{}".format(self.idx_r)] = layer.get(u'yOrigin')

        # Type of SRS
        self.ws_r["J{}".format(self.idx_r)] = layer.get(u'srs_type')
        # EPSG code
        self.ws_r["K{}".format(self.idx_r)] = layer.get(u'EPSG')

        # Creation date
        self.ws_r["M{}".format(self.idx_r)] = layer.get(u'date_crea')
        # Last update date
        self.ws_r["N{}".format(self.idx_r)] = layer.get(u'date_actu')

        # Number of bands
        self.ws_r["O{}".format(self.idx_r)] = layer.get(u'num_bands')

        # Format of data
        self.ws_r["P{}".format(self.idx_r)] = "{0} {1}".format(layer.get(u'format'),
                                               layer.get('format_version'))
        # Compression rate
        self.ws_r["Q{}".format(self.idx_r)] = layer.get(u'compr_rate')

        # Color referential
        self.ws_r["R{}".format(self.idx_r)] = layer.get(u'color_ref')

        # Dependencies
        self.ws_r["S{}".format(self.idx_v)].style.alignment.wrap_text = True
        self.ws_r.cell("S{}".format(self.idx_v)).value = u' |\n '.join(layer.get(u'dependencies'))

        # total size of file and its dependencies
        self.ws_r["T{}".format(self.idx_r)] = layer.get(u'total_size')

        # in case of a source error
        if layer.get('err_gdal')[0] != 0:
            logging.warning('\tproblem detected')
            self.ws_r["U{}".format(self.idx_r)] = "{0} : {1}".format(layer.get('err_gdal')[0],
                                                                     layer.get('err_gdal')[1])
            self.ws_r["U{}".format(self.idx_r)].style = self.s_error
        else:
            pass

        # end of method
        return

    def store_md_fdb(self, filedb):
        """ Storing metadata about a file database
        """
        # increment line
        self.idx_f += 1

        # in case of a source error
        if filedb.get('error'):
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.texts.get(filedb.get('error'))
            logging.warning('\tproblem detected')
            self.ws_fdb["A{}".format(self.idx_f)] = filedb.get('name')
            link = r'=HYPERLINK("{0}","{1}")'.format(filedb.get(u'folder'),
                                                     self.texts.get('browse'))
            self.ws_fdb["B{}".format(self.idx_f)] = link
            self.ws_fdb["B{}".format(self.idx_f)].style = self.s_error
            self.ws_fdb["C{}".format(self.idx_f)] = err_mess
            self.ws_fdb["C{}".format(self.idx_f)].style = self.s_error
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_fdb["A{}".format(self.idx_f)] = filedb.get('name')

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{0}","{1}")'.format(filedb.get(u'folder'),
                                                 self.texts.get('browse'))
        self.ws_fdb["B{}".format(self.idx_f)] = link
        self.ws_fdb["B{}".format(self.idx_f)].style = self.s_link

        self.ws_fdb["C{}".format(self.idx_f)] = path.basename(filedb.get(u'folder'))
        self.ws_fdb["D{}".format(self.idx_f)] = filedb.get(u'total_size')
        self.ws_fdb["E{}".format(self.idx_f)] = filedb.get(u'date_crea')
        self.ws_fdb["F{}".format(self.idx_f)] = filedb.get(u'date_actu')
        self.ws_fdb["G{}".format(self.idx_f)] = filedb.get(u'layers_count')
        self.ws_fdb["H{}".format(self.idx_f)] = filedb.get(u'total_fields')
        self.ws_fdb["I{}".format(self.idx_f)] = filedb.get(u'total_objs')

        # in case of a source error
        if filedb.get('err_gdal')[0] != 0:
            logging.warning('\tproblem detected')
            self.ws_fdb["P{}".format(self.idx_f)] = "{0} : {1}".format(filedb.get('err_gdal')[0],
                                                                       filedb.get('err_gdal')[1])
            self.ws_fdb["P{}".format(self.idx_f)].style = self.s_error
        else:
            pass

        # parsing layers
        for (layer_idx, layer_name) in zip(filedb.get(u'layers_idx'),
                                           filedb.get(u'layers_names')):
            # increment line
            self.idx_f += 1
            champs = ""
            # get the layer informations
            try:
                gdb_layer = filedb.get('{0}_{1}'.format(layer_idx,
                                                        layer_name))
            except UnicodeDecodeError:
                gdb_layer = filedb.get('{0}_{1}'.format(layer_idx,
                                                        unicode(layer_name.decode('latin1'))))
            # in case of a source error
            if gdb_layer.get('error'):
                err_mess = self.texts.get(gdb_layer.get('error'))
                logging.warning('\tproblem detected: \
                                  {0} in {1}'.format(err_mess,
                                                     gdb_layer.get(u'title')))
                sheet.write(line, 6, gdb_layer.get(u'title'), self.xls_erreur)
                sheet.write(line, 7, err_mess, self.xls_erreur)
                # Interruption of function
                continue
            else:
                pass
            # in case of a source error
            if gdb_layer.get('error'):
                # sheet.row(line).set_style(self.xls_erreur)
                err_mess = self.texts.get(gdb_layer.get('error'))
                logging.warning('\tproblem detected')
                self.ws_fdb["A{}".format(self.idx_f)] = gdb_layer.get('title')
                self.ws_fdb["C{}".format(self.idx_f)] = err_mess
                # keep looping
                continue
            else:
                pass

            self.ws_fdb["G{}".format(self.idx_f)] = gdb_layer.get('title')
            self.ws_fdb["H{}".format(self.idx_f)] = gdb_layer.get(u'num_fields')
            self.ws_fdb["I{}".format(self.idx_f)] = gdb_layer.get(u'num_obj')
            self.ws_fdb["J{}".format(self.idx_f)] = gdb_layer.get(u'type_geom')
            self.ws_fdb["K{}".format(self.idx_f)] = gdb_layer.get(u'srs')
            self.ws_fdb["L{}".format(self.idx_f)] = gdb_layer.get(u'srs_type')
            self.ws_fdb["M{}".format(self.idx_f)] = gdb_layer.get(u'EPSG')

            # Spatial extent
            emprise = u"Xmin : {0} - Xmax : {1} | \nYmin : {2} - Ymax : {3}"\
                      .format(unicode(gdb_layer.get(u'Xmin')),
                              unicode(gdb_layer.get(u'Xmax')),
                              unicode(gdb_layer.get(u'Ymin')),
                              unicode(gdb_layer.get(u'Ymax')))
            self.ws_fdb["N{}".format(self.idx_f)].style.alignment.wrap_text = True
            self.ws_fdb["N{}".format(self.idx_f)] = emprise

            # Field informations
            fields = gdb_layer.get(u'fields')
            for chp in fields.keys():
                # field type
                if 'Integer' in fields[chp][0]:
                    tipo = self.texts.get(u'entier')
                elif fields[chp][0] == 'Real':
                    tipo = self.texts.get(u'reel')
                elif fields[chp][0] == 'String':
                    tipo = self.texts.get(u'string')
                elif fields[chp][0] == 'Date':
                    tipo = self.texts.get(u'date')
                else:
                    tipo = "unknown"
                    logging.warning(chp, " unknown type")

                # concatenation of field informations
                try:
                    champs = champs + chp +\
                              u" (" + tipo + self.texts.get(u'longueur') +\
                              unicode(fields[chp][1]) +\
                              self.texts.get(u'precision') +\
                              unicode(fields[chp][2]) + u") ; "
                except UnicodeDecodeError:
                    logging.warning('Field name with special letters: {}'.format(chp.decode('latin1')))
                    # decode the fucking field name
                    champs = champs + chp.decode('latin1') \
                    + u" ({}, Lg. = {}, Pr. = {}) ;".format(tipo,
                                                            fields[chp][1],
                                                            fields[chp][2])
                    # then continue
                    continue

            # Once all fieds explored, write them
            self.ws_fdb["O{}".format(self.idx_f)] = champs

        # end of method
        return

    def store_md_mapdoc(self, mapdoc):
        """ TO DOCUMENT
        """
        # increment line
        self.idx_m += 1

        # local variables
        champs = ""

        # in case of a source error
        if mapdoc.get('error'):
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.texts.get(mapdoc.get('error'))
            logging.warning('\tproblem detected')
            self.ws_mdocs["A{}".format(self.idx_m)] = mapdoc.get('name')
            self.ws_mdocs["A{}".format(self.idx_m)].style = self.s_error
            link = r'=HYPERLINK("{0}","{1}")'.format(mapdoc.get(u'folder'),
                                                     self.texts.get('browse'))
            self.ws_mdocs["B{}".format(self.idx_m)] = link
            self.ws_mdocs["B{}".format(self.idx_m)].style = self.s_error
            self.ws_mdocs["C{}".format(self.idx_m)] = err_mess
            self.ws_mdocs["C{}".format(self.idx_m)].style = self.s_error
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_mdocs["A{}".format(self.idx_m)] = mapdoc.get('name')

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{0}","{1}")'.format(mapdoc.get(u'folder'),
                                                 self.texts.get('browse'))
        self.ws_mdocs["B{}".format(self.idx_m)] = link
        self.ws_mdocs["B{}".format(self.idx_m)].style = self.s_link

        self.ws_mdocs["C{}".format(self.idx_m)] = mapdoc.get('title')
        self.ws_mdocs["D{}".format(self.idx_m)] = mapdoc.get('creator_prod')
        self.ws_mdocs["E{}".format(self.idx_m)] = mapdoc.get('keywords')
        self.ws_mdocs["F{}".format(self.idx_m)] = mapdoc.get('subject')

        # end of method
        return

    def store_md_cad(self, cad):
        """ TO DOCUMENT
        """
        # increment line
        self.idx_c += 1

        # local variables
        champs = ""

        # in case of a source error
        if cad.get('error'):
            # sheet.row(line).set_style(self.xls_erreur)
            err_mess = self.texts.get(cad.get('error'))
            logging.warning('\tproblem detected')
            self.ws_cad["A{}".format(self.idx_c)] = cad.get('name')
            self.ws_cad["A{}".format(self.idx_c)].style = self.s_error
            link = r'=HYPERLINK("{0}","{1}")'.format(cad.get(u'folder'),
                                                     self.texts.get('browse'))
            self.ws_cad["B{}".format(self.idx_c)] = link
            self.ws_cad["B{}".format(self.idx_c)].style = self.s_error
            self.ws_cad["C{}".format(self.idx_c)] = err_mess
            self.ws_cad["C{}".format(self.idx_c)].style = self.s_error
            # Interruption of function
            return False
        else:
            pass

        # Name
        self.ws_cad["A{}".format(self.idx_c)] = cad.get('name')

        # Path of parent folder formatted to be a hyperlink
        link = r'=HYPERLINK("{0}","{1}")'.format(cad.get(u'folder'),
                                                 self.texts.get('browse'))
        self.ws_cad["B{}".format(self.idx_c)] = link
        self.ws_cad["B{}".format(self.idx_c)].style = self.s_link

###############################################################################
###### Stand alone program ########
###################################

if __name__ == '__main__':
    """ Standalone execution and tests
    """
    # ------------ Specific imports ---------------------
    from os import path

    # ------------ REAL START ----------------------------
    wb = files2xlsx()
    # wb.set_worksheets(has_vector=1,
    #                   has_raster=1,
    #                   has_service=1,
    #                   has_resource=1)

    # # saving the test file
    # wb.save(r"files2xlsx.xlsx")
